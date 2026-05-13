from pathlib import Path

import pytest
from cli.core.products import app as product_app
from cli.core.stats import ProductStatsCollector
from mpt_api_client.models import Meta as ClientMeta
from mpt_api_client.models import Model, ModelCollection, Pagination
from openpyxl.reader.excel import load_workbook
from typer.testing import CliRunner

runner = CliRunner()
PRODUCT_ID = "PRD-0232-2541"
EXPECTED_PRODUCT_SHEETS = (
    "General",
    "Settings",
    "Items",
    "Items Groups",
    "Parameters Groups",
    "Agreements Parameters",
    "Assets Parameters",
    "Item Parameters",
    "Request Parameters",
    "Subscription Parameters",
    "Templates",
)
EXPECTED_PRODUCT_MAIN_SHEET_VALUES = (
    ("General", 12, (("A1", "General Information"), ("A2", "Product ID"), ("B2", PRODUCT_ID))),
    (
        "Settings",
        12,
        (("A1", "Setting"), ("A2", "Change order validation (draft)"), ("C2", "Enabled")),
    ),
)
EXPECTED_PRODUCT_ID_ROWS = (
    ("Items", "ITM-0232-2541-0001"),
    ("Items Groups", "IGR-0232-2541-0001"),
    ("Parameters Groups", "PGR-0232-2541-0002"),
    ("Agreements Parameters", "PAR-0232-2541-0001"),
    ("Assets Parameters", "PAR-0232-2541-0027"),
    ("Item Parameters", "PAR-0232-2541-0022"),
    ("Request Parameters", "PAR-0232-2541-0012"),
    ("Subscription Parameters", "PAR-0232-2541-0023"),
    ("Templates", "TPL-0232-2541-0005"),
)


class ModelCollectionFactory:
    def __init__(self, mocker):
        self.mocker = mocker

    def __call__(self, data_list):
        collection = self.mocker.MagicMock(spec=ModelCollection)
        collection.meta = self.mocker.MagicMock(spec=ClientMeta)
        collection.meta.pagination = self.mocker.MagicMock(spec=Pagination)
        collection.meta.pagination.limit = 100
        collection.meta.pagination.offset = 0
        collection.meta.pagination.total = len(data_list)
        collection.resources = []
        for resource_data in data_list:
            resource = self.mocker.MagicMock(spec=Model)
            resource.to_dict.return_value = resource_data
            collection.resources.append(resource)
        return collection


@pytest.fixture
def model_collection_factory(mocker):
    return ModelCollectionFactory(mocker)


@pytest.fixture
def product_catalog(account_container_mock):
    return account_container_mock.api_mpt_client().catalog.products


@pytest.fixture
def catalog_items(account_container_mock):
    return account_container_mock.api_mpt_client().catalog.items


@pytest.fixture
def product_parameters(product_catalog):
    return product_catalog.parameters.return_value


@pytest.fixture
def export_product_resource(mocker, product_catalog, mpt_product_data):
    product_resource = mocker.MagicMock(spec=Model)
    product_resource.to_dict.return_value = mpt_product_data
    product_catalog.get.return_value = product_resource


@pytest.fixture
def export_product_related_resources(
    catalog_items,
    product_catalog,
    product_parameters,
    product_related_data,
    model_collection_factory,
):
    items_select = catalog_items.filter.return_value.select
    item_groups_select = product_catalog.item_groups.return_value.select
    parameter_groups_select = product_catalog.parameter_groups.return_value.select
    parameters_select = product_parameters.filter.return_value.select
    templates_select = product_catalog.templates.return_value.select
    items_select.return_value.fetch_page.return_value = model_collection_factory([
        product_related_data["mpt_item"]
    ])
    item_groups_select.return_value.fetch_page.return_value = model_collection_factory([
        product_related_data["mpt_item_group"]
    ])
    parameter_groups_select.return_value.fetch_page.return_value = model_collection_factory([
        product_related_data["mpt_parameter_group"]
    ])
    parameters_select.return_value.fetch_page.side_effect = [
        model_collection_factory([product_related_data["mpt_agreement_parameter"]]),
        model_collection_factory([product_related_data["mpt_asset_parameter"]]),
        model_collection_factory([product_related_data["mpt_item_parameter"]]),
        model_collection_factory([product_related_data["mpt_request_parameter"]]),
        model_collection_factory([product_related_data["mpt_subscription_parameter"]]),
    ]
    templates_select.return_value.fetch_page.return_value = model_collection_factory([
        product_related_data["mpt_template"]
    ])


@pytest.fixture
def exported_product_workbook(tmp_path, export_product_resource, export_product_related_resources):
    result = runner.invoke(
        product_app, ["export", PRODUCT_ID, "--out", str(tmp_path)], input="y\ny\n"
    )
    product_path = tmp_path / f"{PRODUCT_ID}.xlsx"

    assert result.exit_code == 0, result.stdout
    assert product_path.exists()
    workbook = load_workbook(product_path)
    yield workbook
    workbook.close()


def test_export_skipped_when_user_declines(mocker, product_container_mock):
    mocker.patch("pathlib.Path.exists", return_value=True)

    result = runner.invoke(product_app, ["export", "PRD-1234"], input="n\n")

    assert result.exit_code == 0, result.stdout
    assert "Skipped export for PRD-1234." in result.stdout
    product_container_mock.product_service().export.assert_not_called()


def test_export_product_account_not_allowed(account_container_mock, active_vendor_account):
    account_container_mock.account.override(active_vendor_account)

    result = runner.invoke(product_app, ["export", "fake_id"])

    assert result.exit_code == 4, result.stdout


def test_export_product_error_exporting_product(mocker, product_container_mock):
    product_container_mock.stats.override(mocker.Mock(spec=ProductStatsCollector, has_errors=True))
    mocker.patch("pathlib.Path.unlink")
    mocker.patch("pathlib.Path.replace")

    result = runner.invoke(product_app, ["export", "fake_id"], input="y\n")

    assert result.exit_code == 3, result.stdout


def test_export_product_overwrites_existing_files(mocker, tmp_path, product_container_mock):
    mocker.patch.object(Path, "exists", return_value=True)
    mocker.patch.object(Path, "unlink", return_value=True)
    mocker.patch.object(Path, "replace")
    product_container_mock.stats.override(mocker.Mock(spec=ProductStatsCollector, has_errors=False))
    product_id = "PRD-1234"
    tmp_file = tmp_path / f"{product_id}.xlsx"
    tmp_file.touch()

    result = runner.invoke(
        product_app,
        ["export", product_id, "--out", str(tmp_file)],
        input="y\ny\n",
    )

    assert result.exit_code == 0, result.stdout
    Path.exists.assert_called()
    assert tmp_file.exists()
    Path.unlink.assert_called()
    Path.replace.assert_called_once()
    product_container_mock.product_service().export.assert_called_once()


@pytest.mark.integration
def test_export_product(exported_product_workbook):
    sheet_names = exported_product_workbook.sheetnames  # act

    assert sheet_names == list(EXPECTED_PRODUCT_SHEETS)


@pytest.mark.integration
def test_export_product_main_sheet_values(exported_product_workbook):
    main_sheet_values = EXPECTED_PRODUCT_MAIN_SHEET_VALUES  # act

    for sheet_values in main_sheet_values:
        assert exported_product_workbook[sheet_values[0]].max_row == sheet_values[1]
        for cell_name, expected_value in sheet_values[2]:
            assert exported_product_workbook[sheet_values[0]][cell_name].value == expected_value


@pytest.mark.integration
def test_export_product_related_sheet_values(exported_product_workbook):
    id_rows = EXPECTED_PRODUCT_ID_ROWS  # act

    for id_row in id_rows:
        assert exported_product_workbook[id_row[0]].max_row == 2
        assert exported_product_workbook[id_row[0]]["A1"].value == "ID"
        assert exported_product_workbook[id_row[0]]["A2"].value == id_row[1]
