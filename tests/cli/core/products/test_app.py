from pathlib import Path
from unittest.mock import Mock
from urllib.parse import urljoin

import pytest
from cli.core.models import DataCollectionModel
from cli.core.products import app
from cli.core.products.app import create_product, update_product
from cli.core.services.service_result import ServiceResult
from cli.core.stats import ProductStatsCollector
from openpyxl.reader.excel import load_workbook
from requests import Response
from typer.testing import CliRunner

runner = CliRunner()


def test_list_products(
    expected_account, mocker, requests_mocker, mpt_client, mpt_products_response
):
    mocker.patch("cli.core.products.app.get_active_account", return_value=expected_account)
    requests_mocker.get(
        urljoin(
            mpt_client.base_url,
            "/catalog/products?limit=10&offset=0",
            allow_fragments=True,
        ),
        json=mpt_products_response,
    )

    result = runner.invoke(app, ["list"])

    assert result.exit_code == 0, result.stdout
    assert mpt_products_response["data"][0]["id"] in result.stdout


def test_list_products_with_query_and_paging(
    expected_account, mocker, requests_mocker, mpt_client, mpt_products_response
):
    mocker.patch("cli.core.products.app.get_active_account", return_value=expected_account)
    requests_mocker.get(
        urljoin(
            mpt_client.base_url,
            "/catalog/products?limit=20&offset=0&eq(product.id,PRD-1234)",
            allow_fragments=True,
        ),
        json=mpt_products_response,
    )

    result = runner.invoke(
        app,
        ["list", "--page", "20", "--query", "eq(product.id,PRD-1234)"],
    )

    assert result.exit_code == 0, result.stdout
    assert mpt_products_response["data"][0]["id"] in result.stdout


def test_sync_not_valid_definition(mocker, product_container_mock):
    mocker.patch.object(
        product_container_mock.product_service(),
        "validate_definition",
        return_value=ServiceResult(success=False, model=None, stats=Mock()),
    )

    result = runner.invoke(
        app,
        ["sync", "--dry-run", "some-file"],
    )

    assert result.exit_code == 3, result.stdout


def test_sync_with_dry_run(mocker, product_container_mock):
    create_mock = mocker.patch("cli.core.products.app.create_product")
    update_mock = mocker.patch("cli.core.products.app.update_product")

    result = runner.invoke(
        app,
        ["sync", "--dry-run", "fake_file.xlsx"],
    )

    assert result.exit_code == 0, result.stdout
    assert "Product definition" in result.stdout
    create_mock.assert_not_called()
    update_mock.assert_not_called()


def test_sync_product_update(mocker, product_container_mock, product_data_from_dict):
    validate_definition_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "validate_definition",
        return_value=ServiceResult(success=True, model=None, stats=Mock()),
    )
    retrieve_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "retrieve",
        return_value=ServiceResult(success=True, model=product_data_from_dict, stats=Mock()),
    )
    update_product_mock = mocker.patch("cli.core.products.app.update_product")

    result = runner.invoke(
        app,
        ["sync", "fake_file.xlsx"],
        input="y\n",
    )

    assert result.exit_code == 0, result.stdout
    assert "Do you want to update product" in result.stdout
    validate_definition_mock.assert_called_once()
    retrieve_mock.assert_called_once()
    update_product_mock.assert_called_once()


def test_sync_product_update_error(mocker, product_container_mock, product_data_from_dict):
    mocker.patch.object(ProductStatsCollector, "has_errors", new=True)
    validate_definition_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "validate_definition",
        return_value=ServiceResult(success=True, model=None, stats=Mock()),
    )
    retrieve_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "retrieve",
        return_value=ServiceResult(success=True, model=product_data_from_dict, stats=Mock()),
    )
    update_product_mock = mocker.patch("cli.core.products.app.update_product")

    result = runner.invoke(
        app,
        ["sync", "fake_file.xlsx"],
        input="y\n",
    )

    assert result.exit_code == 3, result.stdout
    validate_definition_mock.assert_called_once()
    retrieve_mock.assert_called_once()
    update_product_mock.assert_called_once()


def test_sync_product_force_create(
    mocker, product_container_mock, product_new_file, product_data_from_dict
):
    mocker.patch.object(ProductStatsCollector, "has_errors", new=False)
    to_table_mock = mocker.patch.object(
        ProductStatsCollector, "to_table", return_value=Mock(return_value="")
    )
    validate_definition_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "validate_definition",
        return_value=ServiceResult(success=True, model=None, stats=Mock()),
    )
    retrieve_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "retrieve",
        return_value=ServiceResult(success=True, model=product_data_from_dict, stats=Mock()),
    )
    create_product_mock = mocker.patch("cli.core.products.app.create_product")

    result = runner.invoke(
        app,
        ["sync", "--force-create", str(product_new_file)],
        input="y\n",
    )

    assert result.exit_code == 0, result.stdout
    assert "Do you want to create new" in result.stdout
    validate_definition_mock.assert_called_once()
    retrieve_mock.assert_called_once()
    create_product_mock.assert_called_once()
    to_table_mock.assert_called_once()


def test_sync_product_no_product(mocker, product_new_file, product_container_mock):
    mocker.patch.object(ProductStatsCollector, "has_errors", new=False)
    to_table_mock = mocker.patch.object(
        ProductStatsCollector, "to_table", return_value=Mock(return_value="")
    )
    validate_definition_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "validate_definition",
        return_value=ServiceResult(success=True, model=None, stats=Mock()),
    )
    retrieve_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "retrieve",
        return_value=ServiceResult(success=True, model=None, stats=Mock()),
    )
    create_mock = mocker.patch("cli.core.products.app.create_product")

    result = runner.invoke(
        app,
        ["sync", str(product_new_file)],
        input="y\n",
    )

    assert result.exit_code == 0, result.stdout
    assert "Do you want to create new product for account" in result.stdout
    validate_definition_mock.assert_called_once()
    retrieve_mock.assert_called_once()
    create_mock.assert_called_once()
    to_table_mock.assert_called_once()


def test_create_product(
    mocker,
    expected_account,
    mpt_client,
    product_new_file,
    product_data_from_dict,
    item_data_from_dict,
    parameter_group_data_from_dict,
    parameters_data_from_dict,
    template_data_from_dict,
    product_container_mock,
):
    create_product_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "create",
        return_value=ServiceResult(success=True, model=product_data_from_dict, stats=Mock()),
    )
    create_item_group_mock = mocker.patch.object(
        product_container_mock.item_group_service(),
        "create",
        return_value=ServiceResult(
            success=True,
            model=None,
            collection=DataCollectionModel(collection={"fake_id": item_data_from_dict}),
            stats=Mock(),
        ),
    )
    create_parameter_group_mock = mocker.patch.object(
        product_container_mock.parameter_group_service(),
        "create",
        return_value=ServiceResult(
            success=True,
            model=None,
            collection=DataCollectionModel(collection={"fake_id": parameters_data_from_dict}),
            stats=Mock(),
        ),
    )
    create_parameter_service_mock = mocker.patch.object(
        product_container_mock.agreement_parameters_service(),
        "create",
        return_value=ServiceResult(
            success=True,
            model=None,
            collection=DataCollectionModel(collection={"fake_id": parameters_data_from_dict}),
            stats=Mock(),
        ),
    )
    set_new_parameter_mock = mocker.patch.object(
        product_container_mock.agreement_parameters_service(), "set_new_parameter_group"
    )
    create_template_service_mock = mocker.patch.object(
        product_container_mock.template_service(),
        "create",
        return_value=ServiceResult(success=True, model=template_data_from_dict, stats=Mock()),
    )
    set_new_parameter_template_mock = mocker.patch.object(
        product_container_mock.template_service(), "set_new_parameter_group"
    )
    create_item_service_mock = mocker.patch.object(
        product_container_mock.item_service(),
        "create",
        return_value=ServiceResult(success=True, model=item_data_from_dict, stats=Mock()),
    )
    set_new_item_groups_mock = mocker.patch.object(
        product_container_mock.item_service(), "set_new_item_groups"
    )
    add_collection_spy = mocker.spy(DataCollectionModel, "add")

    create_product(product_container_mock, Mock())  # act

    create_product_mock.assert_called_once()
    assert product_container_mock.resource_id() == product_data_from_dict.id
    create_item_group_mock.assert_called_once()
    create_parameter_group_mock.assert_called_once()
    create_parameter_service_mock.assert_called_once()
    set_new_parameter_mock.asset_called_once()
    assert add_collection_spy.call_count == 4
    create_template_service_mock.assert_called_once()
    set_new_parameter_template_mock.assert_called_once()
    create_item_service_mock.assert_called_once()
    set_new_item_groups_mock.assert_called_once()


def test_create_product_error(mocker, product_container_mock):
    stats = product_container_mock.stats()
    create_product_mock = mocker.patch.object(
        product_container_mock.product_service(),
        "create",
        return_value=ServiceResult(success=False, model=None, stats=stats),
    )
    create_item_group_spy = mocker.spy(product_container_mock.item_service(), "create")

    create_product(product_container_mock, Mock())  # act

    create_product_mock.assert_called_once()
    create_item_group_spy.assert_not_called()


def test_export_product_account_not_allowed(account_container_mock, active_vendor_account):
    account_container_mock.account.override(active_vendor_account)

    result = runner.invoke(app, ["export", "fake_id"])

    assert result.exit_code == 4, result.stdout


def test_export_product_error_exporting_product(account_container_mock, product_container_mock):
    product_container_mock.stats.override(Mock(ProductStatsCollector, has_errors=True))

    result = runner.invoke(app, ["export", "fake_id"], input="y\n")

    assert result.exit_code == 3, result.stdout


def test_export_product_overwrites_existing_files(
    mocker,
    operations_account,
    tmp_path,
    product_container_mock,
):
    exists_mock = mocker.patch.object(Path, "exists", return_value=True)
    unlink_mock = mocker.patch.object(Path, "unlink", return_value=True)
    product_id = "PRD-1234"
    tmp_file = tmp_path / f"{product_id}.xlsx"
    tmp_file.touch()

    result = runner.invoke(
        app,
        ["export", product_id, "--out", str(tmp_file)],
        input="y\ny\n",
    )

    assert result.exit_code == 0, result.stdout
    exists_mock.assert_called_once()
    assert tmp_file.exists()
    unlink_mock.assert_called_once()
    product_container_mock.product_service().export.assert_called_once()


def test_update_product(product_container_mock):
    update_product(product_container_mock, Mock())  # act

    product_container_mock.item_service().update.assert_called_once()
    product_container_mock.item_group_service().update.assert_called_once()
    product_container_mock.template_service().update.assert_called_once()
    product_container_mock.parameter_group_service().update.assert_called_once()
    product_container_mock.agreement_parameters_service().update.assert_called_once()
    product_container_mock.asset_parameters_service().update.assert_called_once()
    product_container_mock.request_parameters_service().update.assert_called_once()
    product_container_mock.subscription_parameters_service().update.assert_called_once()


@pytest.mark.integration
def test_export_product(
    mocker,
    tmp_path,
    account_container_mock,
    list_response_mock_data_factory,
    mpt_product_data,
    mpt_item_data,
    mpt_item_group_data,
    mpt_parameter_group_data,
    mpt_agreement_parameter_data,
    mpt_asset_parameter_data,
    mpt_item_parameter_data,
    mpt_request_parameter_data,
    mpt_subscription_parameter_data,
    mpt_template_data,
):
    mocker.patch.object(
        account_container_mock.mpt_client(),
        "get",
        side_effect=[
            Mock(spec=Response, json=Mock(return_value=mpt_product_data)),
            list_response_mock_data_factory([mpt_item_data]),
            list_response_mock_data_factory([mpt_item_group_data]),
            list_response_mock_data_factory([mpt_parameter_group_data]),
            list_response_mock_data_factory([mpt_agreement_parameter_data]),
            list_response_mock_data_factory([mpt_asset_parameter_data]),
            list_response_mock_data_factory([mpt_item_parameter_data]),
            list_response_mock_data_factory([mpt_request_parameter_data]),
            list_response_mock_data_factory([mpt_subscription_parameter_data]),
            list_response_mock_data_factory([mpt_template_data]),
        ],
    )
    product_id = "PRD-0232-2541"

    result = runner.invoke(app, ["export", product_id, "--out", str(tmp_path)], input="y\ny\n")

    assert result.exit_code == 0, result.stdout
    product_path = tmp_path / f"{product_id}.xlsx"
    assert product_path.exists()
    wb = load_workbook(product_path)
    expected_sheets = [
        "General",
        "Settings",
        "Items",
        "Items Groups",
        "Parameters Groups",
        "Agreements Parameters",
        "Assets Parameters",
        "Item Parameters",
        "Request Parameters",
        "Subscription Parameters",
        "Templates",
    ]
    assert wb.sheetnames == expected_sheets
    general_sheet = wb["General"]
    assert general_sheet.max_row == 12
    assert general_sheet["A1"].value == "General Information"
    assert general_sheet["A2"].value == "Product ID"
    assert general_sheet["B2"].value == product_id
    settings_sheet = wb["Settings"]
    assert settings_sheet.max_row == 12
    assert settings_sheet["A1"].value == "Setting"
    assert settings_sheet["A2"].value == "Change order validation (draft)"
    assert settings_sheet["C2"].value == "Enabled"
    items_sheet = wb["Items"]
    assert items_sheet.max_row == 2
    assert items_sheet["A1"].value == "ID"
    assert items_sheet["A2"].value == "ITM-0232-2541-0001"
    items_groups_sheet = wb["Items Groups"]
    assert items_groups_sheet.max_row == 2
    assert items_groups_sheet["A1"].value == "ID"
    assert items_groups_sheet["A2"].value == "IGR-0232-2541-0001"
    parameters_groups_sheet = wb["Parameters Groups"]
    assert parameters_groups_sheet.max_row == 2
    assert parameters_groups_sheet["A1"].value == "ID"
    assert parameters_groups_sheet["A2"].value == "PGR-0232-2541-0002"
    agreements_params_sheet = wb["Agreements Parameters"]
    assert agreements_params_sheet.max_row == 2
    assert agreements_params_sheet["A1"].value == "ID"
    assert agreements_params_sheet["A2"].value == "PAR-0232-2541-0001"
    agreements_params_sheet = wb["Assets Parameters"]
    assert agreements_params_sheet.max_row == 2
    assert agreements_params_sheet["A1"].value == "ID"
    assert agreements_params_sheet["A2"].value == "PAR-0232-2541-0027"
    item_params_sheet = wb["Item Parameters"]
    assert item_params_sheet.max_row == 2
    assert item_params_sheet["A1"].value == "ID"
    assert item_params_sheet["A2"].value == "PAR-0232-2541-0022"
    request_params_sheet = wb["Request Parameters"]
    assert request_params_sheet.max_row == 2
    assert request_params_sheet["A1"].value == "ID"
    assert request_params_sheet["A2"].value == "PAR-0232-2541-0012"
    subscription_parameter_sheet = wb["Subscription Parameters"]
    assert subscription_parameter_sheet.max_row == 2
    assert subscription_parameter_sheet["A1"].value == "ID"
    assert subscription_parameter_sheet["A2"].value == "PAR-0232-2541-0023"
    templates_sheet = wb["Templates"]
    assert templates_sheet.max_row == 2
    assert templates_sheet["A1"].value == "ID"
    assert templates_sheet["A2"].value == "TPL-0232-2541-0005"
    wb.close()
