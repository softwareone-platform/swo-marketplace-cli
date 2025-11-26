import re
from pathlib import Path
from unittest.mock import patch

import pytest
from cli.core.handlers.errors import (
    RequiredFieldsError,
    RequiredFieldValuesError,
    RequiredSheetsError,
)
from cli.core.handlers.excel_file_handler import ExcelFileHandler
from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


@pytest.fixture
def workbook():
    workbook = Workbook()
    workbook.remove(workbook["Sheet"])
    vertical_sheet = workbook.create_sheet(title="VerticalSheet")
    horizontal_sheet = workbook.create_sheet(title="HorizontalSheet")

    vertical_sheet["A1"] = "Header1"
    vertical_sheet["B1"] = "Header2"
    vertical_sheet["A2"] = "Field1"
    vertical_sheet["A3"] = "Field2"
    vertical_sheet["A4"] = "EmptyField"
    vertical_sheet["B2"] = "Value1"
    vertical_sheet["B3"] = "Value2"

    horizontal_sheet["A1"] = "Header1"
    horizontal_sheet["B1"] = "Header2"
    horizontal_sheet["C1"] = "Header3"
    horizontal_sheet["A2"] = "Value1"
    horizontal_sheet["B2"] = "Value2"
    horizontal_sheet["C2"] = "Value3"

    yield workbook

    workbook.close()


@pytest.fixture
def excel_file_handler(tmp_path, workbook):
    file_path = tmp_path / "fake_excel_file.xlsx"
    workbook.save(file_path)

    with patch("cli.core.handlers.excel_file_handler.load_workbook", return_value=workbook):
        return ExcelFileHandler(file_path)


@pytest.mark.parametrize(
    ("file_path", "expected_path"),
    [("non_extension_file", "non_extension_file.xlsx"), ("test_file.xlsx", "test_file.xlsx")],
)
def test_normalize_file_path(file_path, expected_path):
    result = ExcelFileHandler.normalize_file_path(file_path)

    assert result == Path(expected_path)


def test_create(tmp_path):
    file_path = tmp_path / "fake_file.xlsx"
    handler = ExcelFileHandler(file_path)

    handler.create()  # act

    assert handler.sheet_names == ["General"]


def test_check_required_sheet(excel_file_handler):
    excel_file_handler.check_required_sheet(("VerticalSheet", "HorizontalSheet"))  # act


def test_check_required_sheet_missing_sheet(excel_file_handler):
    with pytest.raises(RequiredSheetsError):
        excel_file_handler.check_required_sheet(("HorizontalSheet", "MissingSheet"))


def test_check_required_fields_in_vertical_sheet(excel_file_handler):
    excel_file_handler.check_required_fields_in_vertical_sheet(  # act
        "VerticalSheet", ("Field1", "Field2")
    )


def test_check_required_fields_in_vertical_sheet_missing_sheet(excel_file_handler):
    with pytest.raises(RequiredFieldsError):
        excel_file_handler.check_required_fields_in_vertical_sheet(
            "VerticalSheet", ["Field1", "MissingField"]
        )


def test_check_required_field_values_in_vertical_sheet(mocker, excel_file_handler):
    test_get_data_from_vertical_sheet_by_fields_mock = mocker.patch.object(
        excel_file_handler,
        "get_data_from_vertical_sheet",
        return_value={"Field1": {"value": "Value1"}, "Field2": {"value": "Value2"}},
    )

    excel_file_handler.check_required_field_values_in_vertical_sheet(  # act
        "VerticalSheet", ["Field1", "Field2"]
    )

    test_get_data_from_vertical_sheet_by_fields_mock.assert_called_once_with("VerticalSheet")


def test_check_required_field_values_in_vertical_sheet_empty_value(mocker, excel_file_handler):
    test_get_data_from_vertical_sheet_by_fields_mock = mocker.patch.object(
        excel_file_handler,
        "get_data_from_vertical_sheet",
        return_value={"Field1": {"value": "Value1"}, "EmptyField": {"value": None}},
    )

    with pytest.raises(RequiredFieldValuesError) as exc_info:
        excel_file_handler.check_required_field_values_in_vertical_sheet(
            "VerticalSheet", ["Field1", "EmptyField"]
        )

    test_get_data_from_vertical_sheet_by_fields_mock.assert_called_once_with("VerticalSheet")
    assert exc_info.value.details == ["EmptyField"]


def test_check_required_fields_in_horizontal_sheet(excel_file_handler):
    excel_file_handler.check_required_fields_in_horizontal_sheet(  # act
        "HorizontalSheet", ["Header1", "Header2"]
    )


def test_check_required_fields_in_horizontal_sheet_missing_field(excel_file_handler):
    with pytest.raises(RequiredFieldsError) as exc_info:
        excel_file_handler.check_required_fields_in_horizontal_sheet(
            "HorizontalSheet", ["Header2", "MissingHeader"]
        )

    assert exc_info.value.details == ["MissingHeader"]


def test_get_cell_value_by_coordinate(excel_file_handler):
    result = excel_file_handler.get_cell_value_by_coordinate("HorizontalSheet", "A1")

    assert result == "Header1"


def test_get_data_from_horizontal_sheet(excel_file_handler):
    result = list(excel_file_handler.get_data_from_horizontal_sheet("HorizontalSheet"))

    assert len(result[0]) == 3
    assert result[0] == {
        "Header1": {"value": "Value1", "coordinate": "A2"},
        "Header2": {"value": "Value2", "coordinate": "B2"},
        "Header3": {"value": "Value3", "coordinate": "C2"},
    }


def test_get_data_from_horizontal_sheet_by_fields(excel_file_handler):
    result = list(excel_file_handler.get_data_from_horizontal_sheet("HorizontalSheet", ["Header2"]))

    assert len(result[0]) == 1
    assert result[0] == {"Header2": {"value": "Value2", "coordinate": "B2"}}


def test_get_data_from_vertical_sheet(excel_file_handler):
    result = excel_file_handler.get_data_from_vertical_sheet("VerticalSheet")

    assert result["Field1"]["value"] == "Value1"
    assert result["Field1"]["coordinate"] == "B2"
    assert result["Field2"]["value"] == "Value2"
    assert result["Field2"]["coordinate"] == "B3"
    assert result["EmptyField"]["value"] is None
    assert result["EmptyField"]["coordinate"] == "B4"


def test_get_data_from_vertical_sheet_by_fields(excel_file_handler):
    result = excel_file_handler.get_data_from_vertical_sheet("VerticalSheet", ("Field1",))

    assert result == {"Field1": {"value": "Value1", "coordinate": "B2"}}


def test_get_values_for_dynamic_sheet(excel_file_handler):
    fields = ["Header1"]
    patterns = [re.compile(r"Value\d+")]

    result = list(
        excel_file_handler.get_values_for_dynamic_sheet("HorizontalSheet", fields, patterns)
    )

    assert len(result[0]) == 1
    assert result[0] == {"Header1": {"value": "Value1", "coordinate": "A2"}}


def test_write(excel_file_handler):
    test_data = [
        {"VerticalSheet": {"A1": "ValueA1"}},
        {"FakeSheet": {"D1": "ValueD1", "J23": "ValueJ23"}},
    ]

    excel_file_handler.write(test_data)  # act

    assert excel_file_handler._get_worksheet("VerticalSheet")["A1"].value == "ValueA1"  # noqa: SLF001
    assert excel_file_handler._get_worksheet("FakeSheet")["D1"].value == "ValueD1"  # noqa: SLF001
    assert excel_file_handler._get_worksheet("FakeSheet")["J23"].value == "ValueJ23"  # noqa: SLF001


def test_write_cell(excel_file_handler):
    excel_file_handler.write_cell("Sheet1", row=2, col=3, value="FakeValue")  # act

    cell = excel_file_handler._get_worksheet("Sheet1")["C2"]  # noqa: SLF001
    assert cell.value == "FakeValue"
    assert cell.style == "Normal"


def test_write_cell_with_style(excel_file_handler):
    fake_style = NamedStyle("fake_style")

    excel_file_handler.write_cell(  # act
        "Sheet1", row=2, col=3, value="FakeValue", style=fake_style
    )

    assert excel_file_handler._get_worksheet("Sheet1")["C2"].style == "fake_style"  # noqa: SLF001


def test_write_cell_with_data_validation(excel_file_handler):
    fake_data_validation = DataValidation()

    excel_file_handler.write_cell(  # act
        "Sheet1", row=2, col=3, value="FakeValue", data_validation=fake_data_validation
    )

    assert "C2" in fake_data_validation.ranges
