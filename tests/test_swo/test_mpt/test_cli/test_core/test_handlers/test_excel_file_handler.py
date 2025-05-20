import re
from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl.workbook import Workbook
from swo.mpt.cli.core.handlers.errors import (
    RequiredFieldsError,
    RequiredFieldValuesError,
    RequiredSheetsError,
)
from swo.mpt.cli.core.handlers.excel_file_handler import ExcelFileHandler


@pytest.fixture
def workbook():
    workbook = Workbook()
    workbook.remove(workbook["Sheet"])
    vertical_sheet = workbook.create_sheet(title="VerticalSheet")
    horizontal_sheet = workbook.create_sheet(title="HorizontalSheet")

    vertical_sheet["A1"] = "Header1"
    vertical_sheet["B1"] = "Header2"
    vertical_sheet["A2"] = "Field1"
    vertical_sheet["A3"] = "Field2"
    vertical_sheet["A4"] = "EmptyField"
    vertical_sheet["B2"] = "Value1"
    vertical_sheet["B3"] = "Value2"

    horizontal_sheet["A1"] = "Header1"
    horizontal_sheet["B1"] = "Header2"
    horizontal_sheet["C1"] = "Header3"
    horizontal_sheet["A2"] = "Value1"
    horizontal_sheet["B2"] = "Value2"
    horizontal_sheet["C2"] = "Value3"

    yield workbook

    workbook.close()


@pytest.fixture
def excel_file_handler(tmp_path, workbook):
    file_path = tmp_path / "fake_excel_file.xlsx"
    workbook.save(file_path)

    with patch("swo.mpt.cli.core.handlers.excel_file_handler.load_workbook", return_value=workbook):
        handler = ExcelFileHandler(file_path)

    return handler


@pytest.mark.parametrize(
    ("file_path", "expected_path"),
    [("non_extension_file", "non_extension_file.xlsx"), ("test_file.xlsx", "test_file.xlsx")],
)
def test_normalize_file_path(file_path, expected_path):
    normalized = ExcelFileHandler.normalize_file_path(file_path)

    assert normalized == Path(expected_path)


def test_check_required_sheet(excel_file_handler):
    excel_file_handler.check_required_sheet(["VerticalSheet", "HorizontalSheet"])


def test_check_required_sheet_missing_sheet(excel_file_handler):
    with pytest.raises(RequiredSheetsError):
        excel_file_handler.check_required_sheet(["HorizontalSheet", "MissingSheet"])


def test_check_required_fields_in_vertical_sheet(excel_file_handler):
    excel_file_handler.check_required_fields_in_vertical_sheet(
        "VerticalSheet", ["Field1", "Field2"]
    )


def test_check_required_fields_in_vertical_sheet_missing_sheet(excel_file_handler):
    with pytest.raises(RequiredFieldsError):
        excel_file_handler.check_required_fields_in_vertical_sheet(
            "VerticalSheet", ["Field1", "MissingField"]
        )


def test_check_required_field_values_in_vertical_sheet(mocker, excel_file_handler):
    test_get_data_from_vertical_sheet_by_fields_mock = mocker.patch.object(
        excel_file_handler,
        "get_data_from_vertical_sheet",
        return_value={"Field1": {"value": "Value1"}, "Field2": {"value": "Value2"}},
    )
    excel_file_handler.check_required_field_values_in_vertical_sheet(
        "VerticalSheet", ["Field1", "Field2"]
    )

    test_get_data_from_vertical_sheet_by_fields_mock.assert_called_once_with("VerticalSheet")


def test_check_required_field_values_in_vertical_sheet_empty_value(mocker, excel_file_handler):
    test_get_data_from_vertical_sheet_by_fields_mock = mocker.patch.object(
        excel_file_handler,
        "get_data_from_vertical_sheet",
        return_value={"Field1": {"value": "Value1"}, "EmptyField": {"value": None}},
    )
    with pytest.raises(RequiredFieldValuesError) as exc_info:
        excel_file_handler.check_required_field_values_in_vertical_sheet(
            "VerticalSheet", ["Field1", "EmptyField"]
        )

    test_get_data_from_vertical_sheet_by_fields_mock.assert_called_once_with("VerticalSheet")
    assert exc_info.value.details == ["EmptyField"]


def test_check_required_fields_in_horizontal_sheet(excel_file_handler):
    excel_file_handler.check_required_fields_in_horizontal_sheet(
        "HorizontalSheet", ["Header1", "Header2"]
    )


def test_check_required_fields_in_horizontal_sheet_missing_field(excel_file_handler):
    with pytest.raises(RequiredFieldsError) as exc_info:
        excel_file_handler.check_required_fields_in_horizontal_sheet(
            "HorizontalSheet", ["Header2", "MissingHeader"]
        )

    assert exc_info.value.details == ["MissingHeader"]


def test_get_cell_value_by_coordinate(excel_file_handler):
    value = excel_file_handler.get_cell_value_by_coordinate("HorizontalSheet", "A1")

    assert value == "Header1"


def test_get_data_from_horizontal_sheet(excel_file_handler):
    data = list(excel_file_handler.get_data_from_horizontal_sheet("HorizontalSheet"))

    assert len(data[0]) == 3
    assert data[0] == {
        "Header1": {"value": "Value1", "coordinate": "A2"},
        "Header2": {"value": "Value2", "coordinate": "B2"},
        "Header3": {"value": "Value3", "coordinate": "C2"},
    }


def test_get_data_from_horizontal_sheet_by_fields(excel_file_handler):
    data = list(excel_file_handler.get_data_from_horizontal_sheet("HorizontalSheet", ["Header2"]))

    assert len(data[0]) == 1
    assert data[0] == {"Header2": {"value": "Value2", "coordinate": "B2"}}


def test_get_data_from_vertical_sheet(excel_file_handler):
    data = excel_file_handler.get_data_from_vertical_sheet("VerticalSheet")

    assert data["Field1"]["value"] == "Value1"
    assert data["Field1"]["coordinate"] == "B2"
    assert data["Field2"]["value"] == "Value2"
    assert data["Field2"]["coordinate"] == "B3"
    assert data["EmptyField"]["value"] is None
    assert data["EmptyField"]["coordinate"] == "B4"


def test_get_data_from_vertical_sheet_by_fields(excel_file_handler):
    data = excel_file_handler.get_data_from_vertical_sheet("VerticalSheet", ["Field1"])

    assert data == {"Field1": {"value": "Value1", "coordinate": "B2"}}


def test_get_values_for_dynamic_sheet(excel_file_handler):
    fields = ["Header1"]
    patterns = [re.compile(r"Value\d+")]

    data = list(
        excel_file_handler.get_values_for_dynamic_sheet("HorizontalSheet", fields, patterns)
    )

    assert len(data[0]) == 1
    assert data[0] == {"Header1": {"value": "Value1", "coordinate": "A2"}}


def test_write(excel_file_handler):
    test_data = [
        {"VerticalSheet": {"A1": "ValueA1"}},
        {"FakeSheet": {"D1": "ValueD1", "J23": "ValueJ23"}},
    ]

    excel_file_handler.write(test_data)

    assert excel_file_handler._get_worksheet("VerticalSheet")["A1"].value == "ValueA1"
    assert excel_file_handler._get_worksheet("FakeSheet")["D1"].value == "ValueD1"
    assert excel_file_handler._get_worksheet("FakeSheet")["J23"].value == "ValueJ23"
