import re
from pathlib import Path

import pytest
from openpyxl import load_workbook
from swo.mpt.cli.core.console import console
from swo.mpt.cli.core.errors import FileNotExistsError, MPTAPIError
from swo.mpt.cli.core.mpt.models import Item
from swo.mpt.cli.core.products import constants
from swo.mpt.cli.core.products.flows import (
    ProductAction,
    check_file_exists,
    check_product_definition,
    check_product_exists,
    get_definition_file,
    get_values_for_dynamic_table,
    sync_items,
    sync_items_groups,
    sync_parameters,
    sync_parameters_groups,
    sync_product_definition,
    sync_templates,
)
from swo.mpt.cli.core.stats import ErrorMessagesCollector, ProductStatsCollector
from swo.mpt.cli.core.utils import get_values_for_table


@pytest.fixture()
def items():
    # see tests/product_files/PRD-1234-1234-1234-file-update.xlsx file
    return [
        Item(id="ITM-1213-3316-0001", name="Adobe PhotoKios"),
        Item(id="ITM-1213-3316-0002", name="Customer"),
        Item(id="ITM-1213-3316-0003", name="Customer"),
        Item(id="ITM-1213-3316-0005", name="Customer"),
    ]


def test_get_definition_file():
    file_path = "/example/PRD-1234-1234.xlsx"

    assert get_definition_file(file_path) == Path(file_path)


def test_get_definition_file_add_postfix():
    file_path = "/example/PRD-1234-1234"

    assert get_definition_file(file_path) == Path(f"{file_path}.xlsx")


def test_check_file_exists(empty_file):
    assert check_file_exists(empty_file)


def test_check_file_not_exists(tmp_path):
    with pytest.raises(FileNotExistsError):
        check_file_exists("tmp_path")


def test_check_product_definition_not_all_tabs(empty_file):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(empty_file, stats)

    expected_message = """General: Required tab doesn't exist
Settings: Required tab doesn't exist\n"""

    assert not stats.is_empty()
    assert str(stats) == expected_message


def test_check_product_definition_not_all_required_general(product_file_root):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root / "PRD-1234-1234-1234-general-not-all.xlsx", stats
    )

    assert not stats.is_empty()
    assert (
        str(stats)
        == """General: Required field Product Name is not provided
\t\tProduct Website: Value is not provided for the required field. Current value: None"""
    )


def test_check_product_definition_not_all_required_parameter_groups(product_file_root):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root / "PRD-1234-1234-1234-parameter-groups-not-all-columns.xlsx",
        stats,
    )

    assert not stats.is_empty()
    assert str(stats) == "Parameters Groups: Required field Label is not provided\n"


def test_check_product_definition_not_all_required_items_groups(product_file_root):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root / "PRD-1234-1234-1234-items-groups-not-all-columns.xlsx",
        stats,
    )

    assert not stats.is_empty()
    assert str(stats) == "Items Groups: Required field Label is not provided\n"


def test_check_product_definition_not_all_required_agreements_parameters(
    product_file_root,
):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root
        / "PRD-1234-1234-1234-agreements-parameters-not-all-columns.xlsx",
        stats,
    )

    assert not stats.is_empty()
    assert (
        str(stats)
        == "Agreements Parameters: Required field ExternalId is not provided\n"
    )


def test_check_product_definition_not_all_required_item_parameters(
    product_file_root,
):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root / "PRD-1234-1234-1234-item-parameters-not-all-columns.xlsx",
        stats,
    )

    assert not stats.is_empty()
    assert str(stats) == "Item Parameters: Required field ExternalId is not provided\n"


def test_check_product_definition_not_all_required_request_parameters(
    product_file_root,
):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root
        / "PRD-1234-1234-1234-request-parameters-not-all-columns.xlsx",
        stats,
    )

    assert not stats.is_empty()
    assert (
        str(stats) == "Request Parameters: Required field ExternalId is not provided\n"
    )


def test_check_product_definition_not_all_required_subscription_parameters(
    product_file_root,
):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root
        / "PRD-1234-1234-1234-subscription-parameters-not-all-columns.xlsx",
        stats,
    )

    assert not stats.is_empty()
    assert (
        str(stats)
        == "Subscription Parameters: Required field ExternalId is not provided\n"
    )


def test_check_product_definition_not_all_required_items(
    product_file_root,
):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root / "PRD-1234-1234-1234-items-not-all-columns.xlsx",
        stats,
    )

    assert not stats.is_empty()
    assert str(stats) == ("Items: Required field Billing Model is not provided, "
                          "Required field Billing Period is not provided\n")


def test_check_product_definition_not_all_required_templates(
    product_file_root,
):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root / "PRD-1234-1234-1234-templates-not-all-columns.xlsx",
        stats,
    )

    assert not stats.is_empty()
    assert str(stats) == "Templates: Required field Content is not provided\n"


def test_check_product_definition_not_all_required_settings(
    product_file_root,
):
    stats = ErrorMessagesCollector()

    stats = check_product_definition(
        product_file_root / "PRD-1234-1234-1234-settings-not-all-columns.xlsx",
        stats,
    )

    assert not stats.is_empty()
    assert str(stats) == "Settings: Required field Value is not provided\n"


def test_sync_parameters_groups(
    mocker, mpt_client, new_product_file, product, parameter_group
):
    stats = ProductStatsCollector()
    parameter_group_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_parameter_group",
        return_value=parameter_group,
    )

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_PARAMETERS_GROUPS]
    values = get_values_for_table(ws, constants.PARAMETERS_GROUPS_FIELDS)

    _, id_mapping = sync_parameters_groups(
        mpt_client, ws, product, values, stats, console.status("")
    )

    assert id_mapping == {
        "PGR-4944-4118-0002": parameter_group,
        "PGR-4944-4118-0003": parameter_group,
    }
    assert (
        ws["A2"].value,
        ws["A3"].value,
    ) == (
        parameter_group.id,
        parameter_group.id,
    )
    assert parameter_group_mock.mock_calls[0].args == (
        mpt_client,
        product,
        {
            "name": "Agreement",
            "label": "Default Group",
            "description": "Description 1",
            "displayOrder": 10,
            "default": True,
        },
    )
    assert parameter_group_mock.mock_calls[1].args == (
        mpt_client,
        product,
        {
            "name": "Customer",
            "label": "Additional Group",
            "description": "Description 2",
            "displayOrder": 20,
            "default": False,
        },
    )


def test_sync_parameters_groups_exception(
    mocker, mpt_client, new_product_file, product
):
    stats = ProductStatsCollector()
    mocker.patch(
        "swo.mpt.cli.core.products.flows.create_parameter_group",
        side_effect=MPTAPIError("Error", "Error"),
    )

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_PARAMETERS_GROUPS]
    values = get_values_for_table(ws, constants.PARAMETERS_GROUPS_FIELDS)

    sync_parameters_groups(mpt_client, ws, product, values, stats, console.status(""))

    assert ws["J2"].value == "Error with response body Error"


def test_sync_items_groups(mocker, mpt_client, new_product_file, product, item_group):
    stats = ProductStatsCollector()
    item_group_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_item_group", return_value=item_group
    )

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_ITEMS_GROUPS]
    values = get_values_for_table(ws, constants.ITEMS_GROUPS_FIELDS)

    _, id_mapping = sync_items_groups(
        mpt_client, ws, product, values, stats, console.status("")
    )

    assert id_mapping == {
        "IGR-4944-4118-0002": item_group,
        "IGR-4944-4118-0003": item_group,
    }
    assert (
        ws["A2"].value,
        ws["A3"].value,
    ) == (
        item_group.id,
        item_group.id,
    )
    assert item_group_mock.mock_calls[0].args == (
        mpt_client,
        product,
        {
            "name": "Agreement",
            "label": "Default Group",
            "description": "Description 1",
            "displayOrder": 10,
            "default": True,
            "multiple": True,
            "required": True,
        },
    )
    assert item_group_mock.mock_calls[1].args == (
        mpt_client,
        product,
        {
            "name": "Customer",
            "label": "Additional Group",
            "description": "Description 2",
            "displayOrder": 20,
            "default": False,
            "multiple": False,
            "required": True,
        },
    )


def test_sync_items_groups_exception(mocker, mpt_client, new_product_file, product):
    stats = ProductStatsCollector()
    mocker.patch(
        "swo.mpt.cli.core.products.flows.create_item_group",
        side_effect=MPTAPIError("Error", "Error"),
    )

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_ITEMS_GROUPS]
    values = get_values_for_table(ws, constants.ITEMS_GROUPS_FIELDS)

    sync_items_groups(mpt_client, ws, product, values, stats, console.status(""))

    assert ws["L2"].value == "Error with response body Error"


def test_sync_parameters(
    mocker, mpt_client, new_product_file, product, parameter, parameter_group
):
    stats = ProductStatsCollector()
    parameter_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_parameter", return_value=parameter
    )

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_AGREEMENTS_PARAMETERS]
    values = get_values_for_table(ws, constants.PARAMETERS_FIELDS)

    _, id_mapping = sync_parameters(
        "Agreement",
        mpt_client,
        ws,
        product,
        values,
        {
            "PGR-4944-4118-0002": parameter_group,
        },
        stats,
        console.status(""),
    )

    assert id_mapping == {
        "PAR-4944-4118-0001": parameter,
        "PAR-4944-4118-0002": parameter,
    }
    assert (
        ws["A2"].value,
        ws["A3"].value,
    ) == (
        parameter.id,
        parameter.id,
    )
    assert ws["H2"].value == parameter_group.id
    assert parameter_mock.mock_calls[0].args == (
        mpt_client,
        product,
        {
            "name": "Parameter 1",
            "description": "Test parameter 1",
            "scope": "Agreement",
            "phase": "Order",
            "type": "Choice",
            "group": {
                "id": parameter_group.id,
            },
            "options": {
                "optionsList": [
                    {
                        "label": "Create account",
                        "value": "New",
                        "order": 0,
                        "description": "Description 1",
                    },
                    {
                        "label": "Migrate account",
                        "value": "Migrate",
                        "order": 1,
                        "description": "Description 2",
                    },
                ],
                "defaultValue": "New",
                "hintText": "Some hint text",
            },
            "constraints": {
                "hidden": False,
                "unique": None,
                "readonly": False,
                "optional": False,
            },
            "externalId": "external_1",
            "displayOrder": 100,
        },
    )
    assert parameter_mock.mock_calls[1].args == (
        mpt_client,
        product,
        {
            "name": "Parameter 2",
            "description": "Test parameter 2",
            "scope": "Agreement",
            "phase": "Fulfillment",
            "type": "SingleLineText",
            "options": {
                "hintText": "Retry Count",
                "placeholderText": "Retry Count",
            },
            "constraints": {
                "hidden": False,
                "unique": False,
                "readonly": None,
                "optional": False,
            },
            "externalId": "external_2",
            "displayOrder": 200,
        },
    )


def test_sync_parameters_exception(
    mocker, mpt_client, new_product_file, product, parameter_group
):
    stats = ProductStatsCollector()
    mocker.patch(
        "swo.mpt.cli.core.products.flows.create_parameter",
        side_effect=MPTAPIError("Error", "Error"),
    )

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_AGREEMENTS_PARAMETERS]
    values = get_values_for_table(ws, constants.PARAMETERS_FIELDS)

    sync_parameters(
        "Agreement",
        mpt_client,
        ws,
        product,
        values,
        {
            "PGR-4944-4118-0002": parameter_group,
        },
        stats,
        console.status(""),
    )

    assert ws["O2"].value == "Error with response body Error"


def test_sync_item(
    mocker,
    mpt_client,
    new_product_file,
    product,
    parameter,
    another_parameter,
    item,
    item_group,
    uom,
):
    stats = ProductStatsCollector()
    item_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.mpt_create_item", return_value=item
    )
    mocker.patch("swo.mpt.cli.core.products.flows.search_uom_by_name", return_value=uom)

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_ITEMS]
    values = get_values_for_dynamic_table(
        ws, constants.ITEMS_FIELDS, [re.compile(r"Parameter\.*")]
    )

    _, id_mapping = sync_items(
        mpt_client,
        ws,
        product,
        values,
        {
            "IGR-4944-4118-0002": item_group,
        },
        {
            "PAR-0000-0000-00001": parameter,
            "PAR-0000-0000-00002": another_parameter,
        },
        stats,
        console.status(""),
    )

    assert id_mapping == {
        "ITM-4944-4118-0001": item,
        "ITM-4944-4118-0002": item,
    }
    assert (
        ws["A2"].value,
        ws["A3"].value,
    ) == (
        item.id,
        item.id,
    )
    assert (
        ws["K2"].value,
        ws["K3"].value,
    ) == (
        item_group.id,
        item_group.id,
    )
    assert item_mock.mock_calls[0].args == (
        mpt_client,
        {
            "name": "Adobe PhotoKiosk",
            "description": "Description 1",
            "externalIds": {
                "vendor": "65AB123BASD",
            },
            "group": {
                "id": item_group.id,
            },
            "product": {
                "id": product.id,
            },
            "quantityNotApplicable": True,
            "terms": {
                "model": "quantity",
                "commitment": "1y",
                "period": "1m",
            },
            "unit": {
                "id": uom.id,
            },
            "parameters": [
                {
                    "id": parameter.id,
                    "value": "New",
                },
                {
                    "id": another_parameter.id,
                    "value": "test ex 21",
                },
            ],
        },
    )
    assert item_mock.mock_calls[1].args == (
        mpt_client,
        {
            "name": "Customer",
            "description": "Description 2",
            "externalIds": {
                "vendor": "65AB123BASD",
            },
            "group": {
                "id": item_group.id,
            },
            "product": {
                "id": product.id,
            },
            "quantityNotApplicable": False,
            "terms": {
                "model": "one-time",
                "period": "one-time",
            },
            "unit": {
                "id": uom.id,
            },
            "parameters": [
                {
                    "id": parameter.id,
                    "value": "Migrate",
                },
                {
                    "id": another_parameter.id,
                    "value": "test ex 22",
                },
            ],
        },
    )


def test_sync_item_exception(
    mocker,
    mpt_client,
    new_product_file,
    product,
    parameter,
    another_parameter,
    item_group,
    uom,
):
    stats = ProductStatsCollector()
    mocker.patch(
        "swo.mpt.cli.core.products.flows.mpt_create_item",
        side_effect=MPTAPIError("Error", "Error"),
    )
    mocker.patch("swo.mpt.cli.core.products.flows.search_uom_by_name", return_value=uom)

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_ITEMS]
    values = get_values_for_dynamic_table(
        ws, constants.ITEMS_FIELDS, [re.compile(r"Parameter\.*")]
    )

    sync_items(
        mpt_client,
        ws,
        product,
        values,
        {
            "IGR-4944-4118-0002": item_group,
        },
        {
            "PAR-0000-0000-00001": parameter,
            "PAR-0000-0000-00002": another_parameter,
        },
        stats,
        console.status(""),
    )

    assert ws["T2"].value == "Error with response body Error"


def test_sync_template(
    mocker,
    mpt_client,
    new_product_file,
    product,
    template,
    parameter,
):
    stats = ProductStatsCollector()
    template_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_template", return_value=template
    )

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_TEMPLATES]
    values = get_values_for_table(ws, constants.TEMPLATES_FIELDS)

    _, id_mapping = sync_templates(
        mpt_client,
        ws,
        product,
        values,
        {
            "PAR-4944-4118-0001": parameter,
        },
        stats,
        console.status(""),
    )

    assert id_mapping == {
        "TPL-4944-4118-0005": template,
        "TPL-4944-4118-0006": template,
    }
    assert (
        ws["A2"].value,
        ws["A3"].value,
    ) == (
        template.id,
        template.id,
    )
    assert template_mock.mock_calls[0].args == (
        mpt_client,
        product,
        {
            "name": "CLI Order completion",
            "type": "OrderCompleted",
            "default": True,
            "content": "Test content **Azure** {{ " + parameter.id + " }}",
        },
    )
    assert template_mock.mock_calls[1].args == (
        mpt_client,
        product,
        {
            "name": "CLI order quering",
            "type": "OrderQuerying",
            "default": False,
            "content": "Test content **Azure**",
        },
    )


def test_sync_template_exception(
    mocker,
    mpt_client,
    new_product_file,
    product,
    parameter,
):
    stats = ProductStatsCollector()
    mocker.patch(
        "swo.mpt.cli.core.products.flows.create_template",
        side_effect=MPTAPIError("Error", "Error"),
    )

    wb = load_workbook(filename=str(new_product_file))
    ws = wb[constants.TAB_TEMPLATES]
    values = get_values_for_table(ws, constants.TEMPLATES_FIELDS)

    sync_templates(
        mpt_client,
        ws,
        product,
        values,
        {
            "PAR-4944-4118-0001": parameter,
        },
        stats,
        console.status(""),
    )

    assert ws["I2"].value == "Error with response body Error"


def test_sync_product(
    mocker,
    mpt_client,
    new_product_file,
    product,
    parameter_group,
    item,
    item_group,
    uom,
    template,
    parameter,
    another_parameter,
    active_vendor_account,
):
    parameter_group_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_parameter_group",
        return_value=parameter_group,
    )
    item_group_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_item_group", return_value=item_group
    )
    parameter_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_parameter",
        side_effect=[
            parameter,
            another_parameter,
            parameter,
            another_parameter,
            parameter,
            another_parameter,
            parameter,
            another_parameter,
        ],
    )
    item_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.mpt_create_item", return_value=item
    )
    mocker.patch("swo.mpt.cli.core.products.flows.search_uom_by_name", return_value=uom)
    template_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_template", return_value=template
    )
    product_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_product", return_value=product
    )

    stats = ProductStatsCollector()
    sync_product_definition(
        mpt_client,
        new_product_file,
        ProductAction.CREATE,
        active_vendor_account,
        stats,
        console.status(""),
    )

    assert parameter_group_mock.call_count == 2
    assert item_group_mock.call_count == 2
    assert parameter_mock.call_count == 8
    assert item_mock.call_count == 2
    assert template_mock.call_count == 2

    product_mock.assert_called_once_with(
        mpt_client,
        {
            "name": "Adobe Commerce (CLI Test)",
            "shortDescription": "Catalog Description",
            "longDescription": "Product Description",
            "website": "https://example.com",
            "externalIds": None,
            "settings": None,
        },
        {
            "itemSelection": False,
            "orderQueueChanges": False,
            "preValidation": {
                "changeOrderDraft": False,
                "productRequest": False,
                "purchaseOrderDraft": True,
                "purchaseOrderQuery": False,
                "terminationOrder": False,
            },
            "productOrdering": True,
            "productRequests": {"enabled": False, "label": None, "title": None},
        },
        Path.cwd() / Path("swo/mpt/cli/core/products/../icons/fake-icon.png"),
    )

    wb = load_workbook(filename=str(new_product_file))
    assert wb[constants.TAB_GENERAL]["B3"].value == product.id


def test_sync_product_extra_columns(
    mocker,
    mpt_client,
    extra_column_product_file,
    product,
    parameter_group,
    item,
    item_group,
    uom,
    template,
    parameter,
    another_parameter,
    active_vendor_account,
):
    parameter_group_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_parameter_group",
        return_value=parameter_group,
    )
    item_group_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_item_group", return_value=item_group
    )
    parameter_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_parameter",
        side_effect=[
            parameter,
            another_parameter,
            parameter,
            another_parameter,
            parameter,
            another_parameter,
            parameter,
            another_parameter,
        ],
    )
    item_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.mpt_create_item", return_value=item
    )
    mocker.patch("swo.mpt.cli.core.products.flows.search_uom_by_name", return_value=uom)
    template_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_template", return_value=template
    )
    product_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.create_product", return_value=product
    )

    stats = ProductStatsCollector()
    sync_product_definition(
        mpt_client,
        extra_column_product_file,
        ProductAction.CREATE,
        active_vendor_account,
        stats,
        console.status(""),
    )

    assert parameter_group_mock.call_count == 2
    assert item_group_mock.call_count == 2
    assert parameter_mock.call_count == 8
    assert item_mock.call_count == 2
    assert template_mock.call_count == 2

    product_mock.assert_called_once_with(
        mpt_client,
        {
            "name": "Adobe Commerce (CLI Test)",
            "shortDescription": "Catalog Description",
            "longDescription": "Product Description",
            "website": "https://example.com",
            "externalIds": None,
            "settings": None,
        },
        {
            "itemSelection": False,
            "orderQueueChanges": False,
            "preValidation": {
                "changeOrderDraft": False,
                "productRequest": False,
                "purchaseOrderDraft": True,
                "purchaseOrderQuery": False,
                "terminationOrder": False,
            },
            "productOrdering": True,
            "productRequests": {"enabled": False, "label": None, "title": None},
        },
        Path.cwd() / Path("swo/mpt/cli/core/products/../icons/fake-icon.png"),
    )

    wb = load_workbook(filename=str(extra_column_product_file))
    assert wb[constants.TAB_GENERAL]["B15"].value == product.id


def test_sync_product_exception(
    mocker, mpt_client, new_product_file, active_vendor_account
):
    mocker.patch(
        "swo.mpt.cli.core.products.flows.create_product",
        side_effect=MPTAPIError("Error", "Error"),
    )

    stats = ProductStatsCollector()
    sync_product_definition(
        mpt_client,
        new_product_file,
        ProductAction.CREATE,
        active_vendor_account,
        stats,
        console.status(""),
    )

    wb = load_workbook(filename=str(new_product_file))
    assert wb[constants.TAB_GENERAL]["C3"].value == "Error with response body Error"


def test_sync_product_update_product(
    mocker,
    mpt_client,
    new_update_product_file,
    active_vendor_account,
    uom,
    items,
):
    review_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.review_item",
        return_value=None,
    )
    publish_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.publish_item",
        return_value=None,
    )
    update_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.mpt_update_item",
        return_value=None,
    )
    unpublish_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.unpublish_item",
        return_value=None,
    )
    get_item_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.get_item",
        side_effect=items,
    )
    create_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.mpt_create_item",
        side_effect=items,
    )
    mocker.patch("swo.mpt.cli.core.products.flows.search_uom_by_name", return_value=uom)

    stats = ProductStatsCollector()
    stats, _ = sync_product_definition(
        mpt_client,
        new_update_product_file,
        ProductAction.UPDATE,
        active_vendor_account,
        stats,
        console.status(""),
    )

    review_mock.assert_called_once_with(mpt_client, "ITM-1213-3316-0001")
    publish_mock.assert_called_once_with(mpt_client, "ITM-1213-3316-0002")
    unpublish_mock.assert_called_once_with(mpt_client, "ITM-1213-3316-0005")
    update_mock.assert_called_once_with(
        mpt_client,
        "ITM-1213-3316-0003",
        {
            "description": "Description 2",
            "externalIds": {"vendor": "65AB123BASD"},
            "group": {"id": "IGR-1213-3316-0002"},
            "name": "Customer",
            "parameters": [],
            "product": {"id": "PRD-1213-3316"},
            "quantityNotApplicable": False,
            "terms": {"model": "usage", "commitment": "1y", "period": "1m"},
            "unit": {"id": "UNT-1916"},
        },
    )
    create_mock.assert_called_once_with(
        mpt_client,
        {
            "name": "Customer",
            "description": "Description 6",
            "group": {"id": "IGR-1213-3316-0002"},
            "product": {"id": "PRD-1213-3316"},
            "quantityNotApplicable": False,
            "terms": {"model": "usage", "commitment": "1y", "period": "1m"},
            "unit": {"id": "UOM-1234-1234"},
            "parameters": [],
            "externalIds": {"vendor": "65AB123BASD2"},
        },
    )
    assert get_item_mock.call_count == 4
    assert stats.tabs[constants.TAB_ITEMS]["skipped"] == 1


def test_sync_product_update_product_operations(
    mocker,
    mpt_client,
    new_update_product_file,
    active_operations_account,
    uom,
    items,
):
    review_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.review_item",
        return_value=None,
    )
    publish_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.publish_item",
        return_value=None,
    )
    update_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.mpt_update_item",
        return_value=None,
    )
    unpublish_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.unpublish_item",
        return_value=None,
    )
    get_item_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.get_item",
        side_effect=items,
    )
    create_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.mpt_create_item",
        side_effect=items,
    )
    mocker.patch("swo.mpt.cli.core.products.flows.search_uom_by_name", return_value=uom)

    stats = ProductStatsCollector()
    stats, _ = sync_product_definition(
        mpt_client,
        new_update_product_file,
        ProductAction.UPDATE,
        active_operations_account,
        stats,
        console.status(""),
    )

    review_mock.assert_called_once_with(mpt_client, "ITM-1213-3316-0001")
    publish_mock.assert_called_once_with(mpt_client, "ITM-1213-3316-0002")
    unpublish_mock.assert_called_once_with(mpt_client, "ITM-1213-3316-0005")
    update_mock.assert_called_once_with(
        mpt_client,
        "ITM-1213-3316-0003",
        {
            "description": "Description 2",
            "externalIds": {
                "operations": "NAV12345",
            },
            "group": {"id": "IGR-1213-3316-0002"},
            "name": "Customer",
            "parameters": [],
            "product": {"id": "PRD-1213-3316"},
            "quantityNotApplicable": False,
            "terms": {"model": "usage", "commitment": "1y", "period": "1m"},
            "unit": {"id": "UNT-1916"},
        },
    )
    create_mock.assert_called_once_with(
        mpt_client,
        {
            "name": "Customer",
            "description": "Description 6",
            "group": {"id": "IGR-1213-3316-0002"},
            "product": {"id": "PRD-1213-3316"},
            "quantityNotApplicable": False,
            "terms": {"model": "usage", "commitment": "1y", "period": "1m"},
            "unit": {"id": "UOM-1234-1234"},
            "parameters": [],
            "externalIds": {"operations": "NAV123456"},
        },
    )
    assert get_item_mock.call_count == 4
    assert stats.tabs[constants.TAB_ITEMS]["skipped"] == 1


def test_check_product_exists(mocker, mpt_client, new_product_file, product):
    get_products_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.get_products",
        return_value=({}, [product]),
    )

    checked_product = check_product_exists(mpt_client, new_product_file)

    assert checked_product == product
    get_products_mock.assert_called_once_with(
        mpt_client, 1, 0, query="id=PRD-1234-1234-1234"
    )


def test_check_product_exists_no_product(mocker, mpt_client, new_product_file):
    get_products_mock = mocker.patch(
        "swo.mpt.cli.core.products.flows.get_products",
        return_value=({}, []),
    )

    checked_product = check_product_exists(mpt_client, new_product_file)

    assert not checked_product
    get_products_mock.assert_called_once_with(
        mpt_client, 1, 0, query="id=PRD-1234-1234-1234"
    )
