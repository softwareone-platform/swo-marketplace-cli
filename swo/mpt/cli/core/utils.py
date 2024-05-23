import re
from typing import Any, Generator, Pattern, TypeAlias

from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.utils.cell import coordinate_from_string  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
from swo.mpt.cli.core.stats import StatsCollector

SheetValue: TypeAlias = tuple[str, str, Any]
SheetValueGenerator: TypeAlias = Generator[list[SheetValue], None, None]


# TODO: add typehints here
def find_first(func, iterable, default=None):
    return next(filter(func, iterable), default)


def get_values_for_general(ws: Worksheet, fields: list[str]) -> list[SheetValue]:
    values: list[SheetValue] = []

    for row in ws.iter_rows(
        min_row=ws.min_row,
        max_row=ws.max_row,
        min_col=0,
        max_col=2,
    ):
        if row[0].value in fields:
            values.append(
                (f"{row[1].column_letter}{row[1].row}", row[0].value, row[1].value)
            )

    return values


def _get_columns_map(ws: Worksheet, fields: list[str]) -> dict[int, str]:
    column_map = {}
    for index, column in enumerate(ws["1"]):
        if column.value and column.value in fields:
            column_map[index] = column.value

    return column_map


def get_values_for_table(ws: Worksheet, fields: list[str]) -> SheetValueGenerator:
    column_map = _get_columns_map(ws, fields)

    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
    ):
        if all(c.value is None for c in row):
            continue

        yield [
            (
                f"{row[index].column_letter}{row[index].row}",
                column_name,
                row[index].value,
            )
            for index, column_name in column_map.items()
        ]


def _get_dynamic_columns_map(
    ws: Worksheet, fields: list[str], patterns: list[Pattern[str]]
) -> dict[int, str]:
    column_map = {}
    for index, column in enumerate(ws["1"]):
        if column.value and (
            column.value in fields or any(re.match(p, column.value) for p in patterns)
        ):
            column_map[index] = column.value

    return column_map


def get_values_for_dynamic_table(
    ws: Worksheet, fields: list[str], patterns: list[Pattern[str]]
) -> SheetValueGenerator:
    column_map = _get_dynamic_columns_map(ws, fields, patterns)

    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
    ):
        if all(c.value is None for c in row):
            continue

        yield [
            (
                f"{row[index].column_letter}{row[index].row}",
                column_name,
                row[index].value,
            )
            for index, column_name in column_map.items()
        ]


def find_value_for(field: str, values: list[SheetValue]) -> SheetValue:
    return find_first(lambda f: f[1] == field, values)


def find_values_by_pattern(
    pattern: Pattern[str], values: list[SheetValue]
) -> list[SheetValue]:
    return list(filter(lambda sv: re.match(pattern, sv[1]), values))


def set_dict_value(original_dict: dict, path: str, value: Any) -> dict:
    path_list = path.split(".", 1)

    if len(path_list) == 1:
        original_dict[path_list[0]] = value
    else:
        current, next_path = path_list[0], path_list[-1]

        if current in original_dict:
            original_dict[current] = set_dict_value(
                original_dict[current], next_path, value
            )
        else:
            original_dict[current] = set_dict_value({}, next_path, value)

    return original_dict


def set_value(
    column_name: str, values: list[SheetValue], value: str
) -> list[SheetValue]:
    return [
        (
            index,
            column,
            value if column == column_name else old_value,
        )
        for index, column, old_value in values
    ]


def add_or_create_error(
    ws: Worksheet, sheet_value: list[SheetValue], exception: Exception
) -> Worksheet:
    column = find_first(
        lambda c: c[1].value == "Error",
        enumerate(ws["1"]),
    )

    if column:
        index, _ = column
        column_letter = get_column_letter(index + 1)
    else:
        column_letter = get_column_letter(ws.max_column + 1)
        ws[f"{column_letter}1"] = "Error"

    index, _, _ = sheet_value[0]
    row_number = coordinate_from_string(index)[1]
    ws[f"{column_letter}{row_number}"] = str(exception)

    return ws


def status_step_text(stats: StatsCollector, tab_name: str) -> str:
    results = stats.tabs[tab_name]

    return (
        f"[green]{results['synced']}[/green] / "
        f"[red bold]{results['error']}[/red bold] / "
        f"[white bold]{results['skipped']}[/white bold] / "
        f"[blue]{results['total']}[/blue]"
    )
