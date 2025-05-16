import enum
import os
import re
from functools import partial
from pathlib import Path
from typing import Optional, TypeAlias, TypeVar

from openpyxl.worksheet.worksheet import Worksheet
from rich.status import Status
from swo.mpt.cli.core.accounts.models import Account
from swo.mpt.cli.core.errors import FileNotExistsError
from swo.mpt.cli.core.handlers.errors import (
    RequiredFieldsError,
    RequiredFieldValuesError,
    RequiredSheetsError,
)
from swo.mpt.cli.core.handlers.excel_file_handler import ExcelFileHandler
from swo.mpt.cli.core.mpt.client import MPTClient
from swo.mpt.cli.core.mpt.flows import (
    create_item as mpt_create_item,
)
from swo.mpt.cli.core.mpt.flows import (
    create_item_group,
    create_parameter,
    create_parameter_group,
    create_product,
    create_template,
    get_item,
    get_products,
    publish_item,
    review_item,
    search_uom_by_name,
    unpublish_item,
)
from swo.mpt.cli.core.mpt.flows import (
    update_item as mpt_update_item,
)
from swo.mpt.cli.core.mpt.models import (
    Item,
    ItemGroup,
    Parameter,
    ParameterGroup,
    Product,
    Template,
)
from swo.mpt.cli.core.products import constants
from swo.mpt.cli.core.products.to_json import (
    to_item_group_json,
    to_item_sync_json,
    to_item_update_or_create_json,
    to_parameter_group_json,
    to_parameter_json,
    to_product_json,
    to_settings_json,
    to_template_json,
)
from swo.mpt.cli.core.stats import ErrorMessagesCollector, ProductStatsCollector
from swo.mpt.cli.core.utils import (
    SheetValue,
    SheetValueGenerator,
    add_or_create_error,
    find_value_for,
    set_value,
    status_step_text,
)

T = TypeVar("T")
SyncResult: TypeAlias = tuple[Worksheet, dict[str, T]]


class ProductAction(enum.Enum):
    CREATE = "create"
    UPDATE = "update"


class ItemAction(enum.Enum):
    CREATE = "create"
    UPDATE = "update"
    REVIEW = "review"
    PUBLISH = "publish"
    UNPUBLISH = "unpublish"
    SKIP = "-"
    SKIPPED = ""


def get_definition_file(path: str) -> Path:
    """
    Returns product definition file path. If only product id is passed assumed
    that product definition file is in the same folder with .xlsx file extension
    """
    return ExcelFileHandler.normalize_file_path(path)


def check_file_exists(product_file_path: Path) -> bool:
    """
    Check that product file exists
    """
    if not ExcelFileHandler(product_file_path).exists():
        raise FileNotExistsError(product_file_path)

    return True


def check_product_definition(
    definition_path: Path, stats: ErrorMessagesCollector
) -> ErrorMessagesCollector:
    """
    Parses Product definition file and check consistency of product definition file
    """
    # check parameters and items refer to proper groups
    file_handler = ExcelFileHandler(definition_path)
    try:
        file_handler.check_required_sheet(constants.REQUIRED_TABS)
    except RequiredSheetsError as error:
        for sheet_name in error.details:
            stats.add_msg(sheet_name, "", "Required tab doesn't exist")

    existing_sheets = set(constants.ALL_TABS).intersection(set(file_handler.sheet_names))
    for sheet_name in existing_sheets:
        if sheet_name not in constants.REQUIRED_FIELDS_BY_TAB:
            continue

        if sheet_name == constants.TAB_GENERAL:
            check_required_general_fields(
                file_handler,
                stats,
                sheet_name,
                constants.REQUIRED_FIELDS_BY_TAB[sheet_name],
                constants.REQUIRED_FIELDS_WITH_VALUES_BY_TAB[sheet_name],
            )
        else:
            check_required_columns(
                file_handler,
                stats,
                sheet_name,
                constants.REQUIRED_FIELDS_BY_TAB[sheet_name],
            )

    return stats


def check_required_general_fields(
    file_handler: ExcelFileHandler,
    stats: ErrorMessagesCollector,
    sheet_name: str,
    required_field_names: list[str],
    required_values_field_names: list[str],
) -> ErrorMessagesCollector:
    """
    Check that required fields and values are presented in General worksheet
    """
    try:
        file_handler.check_required_fields_in_vertical_sheet(sheet_name, required_field_names)
    except RequiredFieldsError as error:
        for field in error.details:
            stats.add_msg(sheet_name,"",f"Required field {field} is not provided")

    try:
        file_handler.check_required_field_values_in_vertical_sheet(
            sheet_name, required_values_field_names
        )
    except RequiredFieldValuesError as error:
        for field in error.details:
            stats.add_msg(
                sheet_name,
                field,
                "Value is not provided for the required field. Current value: None",
            )

    return stats


def check_required_columns(
    file_handler: ExcelFileHandler,
    stats: ErrorMessagesCollector,
    sheet_name: str,
    required_field_names: list[str],
) -> ErrorMessagesCollector:
    """
    Check that required fields and values are presented in tables worksheet
    """
    try:
        file_handler.check_required_fields_in_horizontal_sheet(sheet_name, required_field_names)
    except RequiredFieldsError as error:
        for field in error.details:
            stats.add_msg(
                sheet_name,
                "",
                f"Required field {field} is not provided",
            )
    return stats


def sync_product_definition(
    mpt_client: MPTClient,
    definition_path: Path,
    action: ProductAction,
    active_account: Account,
    stats: ProductStatsCollector,
    status: Status,
) -> tuple[ProductStatsCollector, Optional[Product]]:
    """
    Sync product definition to the marketplace platform
    """
    file_handler = ExcelFileHandler(file_path=definition_path)

    if action == ProductAction.UPDATE:
        stats, product = update_product_definition(
            mpt_client, file_handler, active_account, stats, status
        )
    else:
        stats, product = create_product_definition(mpt_client, file_handler, stats, status)

    file_handler.save()

    return stats, product


def create_product_definition(
    mpt_client: MPTClient,
    file_handler: ExcelFileHandler,
    stats: ProductStatsCollector,
    status: Status,
) -> tuple[ProductStatsCollector, Optional[Product]]:
    general_data = file_handler.get_data_from_vertical_sheet(
        constants.TAB_GENERAL, constants.GENERAL_FIELDS
    )
    settings_data = file_handler.get_data_from_horizontal_sheet(
        constants.TAB_SETTINGS, constants.SETTINGS_FIELDS
    )
    try:
        product = create_product(
            mpt_client,
            to_product_json(general_data),
            to_settings_json(settings_data, constants.SETTINGS_API_MAPPING),
            Path(os.path.dirname(__file__)) / "../icons/fake-icon.png",
        )
        product_data = general_data[constants.GENERAL_PRODUCT_ID]
        file_handler.write([{constants.TAB_GENERAL: {product_data["coordinate"]: product.id}}])
        stats.add_synced(constants.TAB_GENERAL)

    except Exception as e:
        data = [(cell["coordinate"], cell["value"], None) for _, cell in general_data.items()]
        # TODO: re-think the error handling to avoid use wb
        wb = file_handler._workbook
        add_or_create_error(wb[constants.TAB_GENERAL], data, e)
        stats.add_error(constants.TAB_GENERAL)
        return stats, None

    parameter_groups = file_handler.get_data_from_horizontal_sheet(
        constants.TAB_PARAMETERS_GROUPS, constants.PARAMETERS_GROUPS_FIELDS
    )
    _, parameters_groups_id_mapping = sync_parameters_groups(
        mpt_client,
        file_handler,
        product,
        parameter_groups,
        stats,
        status,
    )

    item_groups = file_handler.get_data_from_horizontal_sheet(
        constants.TAB_ITEMS_GROUPS, constants.ITEMS_GROUPS_FIELDS
    )
    _, items_groups_id_mapping = sync_items_groups(
        mpt_client, file_handler, product, item_groups, stats, status
    )

    agreements_parameters = file_handler.get_data_from_horizontal_sheet(
        constants.TAB_AGREEMENT_PARAMETERS, constants.PARAMETERS_FIELDS
    )
    _, agreements_parameters_id_mapping = sync_agreement_parameters(
        mpt_client,
        file_handler,
        product,
        agreements_parameters,
        parameters_groups_id_mapping,
        stats,
        status,
    )

    item_parameters = file_handler.get_data_from_horizontal_sheet(
        constants.TAB_ITEM_PARAMETERS, constants.PARAMETERS_FIELDS
    )
    _, item_parameters_id_mapping = sync_item_parameters(
        mpt_client,
        file_handler,
        product,
        item_parameters,
        parameters_groups_id_mapping,
        stats,
        status,
    )

    request_parameters = file_handler.get_data_from_horizontal_sheet(
        constants.TAB_REQUEST_PARAMETERS, constants.PARAMETERS_FIELDS
    )
    _, request_parameters_id_mapping = sync_request_parameters(
        mpt_client,
        file_handler,
        product,
        request_parameters,
        parameters_groups_id_mapping,
        stats,
        status,
    )

    subscription_parameters = file_handler.get_data_from_horizontal_sheet(
        constants.TAB_SUBSCRIPTION_PARAMETERS, constants.PARAMETERS_FIELDS
    )
    _, subscription_parameters_id_mapping = sync_subscription_parameters(
        mpt_client,
        file_handler,
        product,
        subscription_parameters,
        parameters_groups_id_mapping,
        stats,
        status,
    )

    items = file_handler.get_values_for_dynamic_sheet(
        constants.TAB_ITEMS, constants.ITEMS_FIELDS, [re.compile(r"Parameter\.*")]
    )
    _, items_id_mapping = sync_items(
        mpt_client,
        file_handler,
        product,
        items,
        items_groups_id_mapping,
        item_parameters_id_mapping,
        stats,
        status,
    )

    all_parameters_id_mapping = {
        **agreements_parameters_id_mapping,
        **item_parameters_id_mapping,
        **request_parameters_id_mapping,
        **subscription_parameters_id_mapping,
    }

    templates = file_handler.get_data_from_horizontal_sheet(
        constants.TAB_TEMPLATES, constants.TEMPLATES_FIELDS
    )
    _, templates_id_mapping = sync_templates(
        mpt_client,
        file_handler,
        product,
        templates,
        all_parameters_id_mapping,
        stats,
        status,
    )

    return stats, product


def update_product_definition(
    mpt_client: MPTClient,
    file_handler: ExcelFileHandler,
    active_account: Account,
    stats: ProductStatsCollector,
    status: Status,
) -> tuple[ProductStatsCollector, Optional[Product]]:
    general_values = file_handler.get_data_from_vertical_sheet(
        constants.TAB_GENERAL, constants.GENERAL_FIELDS
    )
    product_id = general_values[constants.GENERAL_PRODUCT_ID]["value"]
    items = file_handler.get_values_for_dynamic_sheet(
        constants.TAB_ITEMS, constants.ITEMS_FIELDS, [re.compile(r"Parameter\.*")]
    )
    update_items(
        mpt_client,
        file_handler,
        product_id,
        items,
        active_account,
        stats,
        status,
    )

    return stats, None


def update_items(
    mpt_client: MPTClient,
    file_handler: ExcelFileHandler,
    product_id: str,
    values: SheetValueGenerator,
    active_account: Account,
    stats: ProductStatsCollector,
    status: Status,
) -> None:
    ws = file_handler._get_worksheet(constants.TAB_ITEMS)
    for sheet_value in values:
        try:
            action = ItemAction(find_value_for(constants.ITEMS_ACTION, sheet_value)[2])
            if action not in (ItemAction.SKIP, ItemAction.SKIPPED, ItemAction.CREATE):
                item_vendor_id = find_value_for(
                    constants.ITEMS_VENDOR_ITEM_ID, sheet_value
                )[2]
                item = get_item(mpt_client, product_id, item_vendor_id)

            match ItemAction(action):
                case ItemAction.REVIEW:
                    review_item(mpt_client, item.id)
                    stats.add_synced(ws.title)
                case ItemAction.PUBLISH:
                    publish_item(mpt_client, item.id)
                    stats.add_synced(ws.title)
                case ItemAction.UNPUBLISH:
                    unpublish_item(mpt_client, item.id)
                    stats.add_synced(ws.title)
                case ItemAction.UPDATE:
                    update_item(
                        mpt_client, active_account, item.id, product_id, sheet_value
                    )
                    stats.add_synced(ws.title)
                case ItemAction.CREATE:
                    create_item(mpt_client, active_account, product_id, ws, sheet_value)
                    stats.add_synced(ws.title)
                case _:
                    stats.add_skipped(ws.title)
        except Exception as e:
            add_or_create_error(ws, sheet_value, e)
            stats.add_error(ws.title)
        finally:
            step_text = status_step_text(stats, ws.title)
            status.update(f"Syncing {ws.title}: {step_text}")


def update_item(
    mpt_client: MPTClient,
    active_account: Account,
    item_id: str,
    product_id: str,
    sheet_value: list[SheetValue],
) -> None:
    item_json = to_item_update_or_create_json(
        product_id, sheet_value, active_account.type == "Operations"
    )
    mpt_update_item(mpt_client, item_id, item_json)


def sync_parameters_groups(
    mpt_client: MPTClient,
    file_handler: ExcelFileHandler,
    product: Product,
    values: SheetValueGenerator,
    stats: ProductStatsCollector,
    status: Status,
) -> SyncResult[ParameterGroup]:
    """
    Sync product parameters groups
    Returns mapping if ids {"<excel-id>": ParameterGroup}
    """
    id_mapping = {}
    ws = file_handler._get_worksheet(constants.TAB_PARAMETERS_GROUPS)
    for sheet_value in values:
        try:
            parameter_group = create_parameter_group(
                mpt_client,
                product,
                to_parameter_group_json(sheet_value),
            )

            index, _, sheet_id_value = find_value_for(
                constants.PARAMETERS_GROUPS_ID, sheet_value
            )

            # TODO: refactor, should be done out of this function
            ws[index] = parameter_group.id
            id_mapping[sheet_id_value] = parameter_group
            stats.add_synced(ws.title)
        except Exception as e:
            add_or_create_error(ws, sheet_value, e)
            stats.add_error(ws.title)
        finally:
            step_text = status_step_text(stats, ws.title)
            status.update(f"Syncing {ws.title}: {step_text}")

    return ws, id_mapping


def sync_items_groups(
    mpt_client: MPTClient,
    file_handler: ExcelFileHandler,
    product: Product,
    values: SheetValueGenerator,
    stats: ProductStatsCollector,
    status: Status,
) -> SyncResult[ItemGroup]:
    """
    Sync item parameters groups
    Returns mapping if ids {"<excel-id>": ItemGroup}
    """
    id_mapping = {}
    ws = file_handler._get_worksheet(constants.TAB_ITEMS_GROUPS)

    for sheet_value in values:
        try:
            item_group = create_item_group(
                mpt_client,
                product,
                to_item_group_json(sheet_value),
            )

            index, _, sheet_id_value = find_value_for(
                constants.ITEMS_GROUPS_ID, sheet_value
            )

            # TODO: refactor, should be done out of this function
            ws[index] = item_group.id
            id_mapping[sheet_id_value] = item_group
            stats.add_synced(ws.title)
        except Exception as e:
            add_or_create_error(ws, sheet_value, e)
            stats.add_error(ws.title)
        finally:
            step_text = status_step_text(stats, ws.title)
            status.update(f"Syncing {ws.title}: {step_text}")

    return ws, id_mapping


def sync_parameters(
    scope: str,
    mpt_client: MPTClient,
    file_handler: ExcelFileHandler,
    product: Product,
    values: SheetValueGenerator,
    parameter_groups_mapping: dict[str, ParameterGroup],
    stats: ProductStatsCollector,
    status: Status,
) -> SyncResult[Parameter]:
    """
    Sync parameters by scope
    Returns mapping if ids {"<excel-id>": Parameter}
    """
    id_mapping = {}
    ws = file_handler._get_worksheet(getattr(constants, f"TAB_{scope.upper()}_PARAMETERS"))
    for sheet_value in values:
        try:
            parameter = create_parameter(
                mpt_client,
                product,
                to_parameter_json(scope, parameter_groups_mapping, sheet_value),
            )

            id_index, _, sheet_id_value = find_value_for(
                constants.ID_COLUMN_NAME, sheet_value
            )

            # TODO: refactor, should be done out of this function
            ws[id_index] = parameter.id
            id_mapping[sheet_id_value] = parameter

            phase = find_value_for(constants.PARAMETERS_PHASE, sheet_value)[2]
            if phase == "Order" and scope not in ("Item", "Request"):
                group_id_index, _, sheet_id_value = find_value_for(
                    constants.PARAMETERS_GROUP_ID, sheet_value
                )
                created_group = parameter_groups_mapping[sheet_id_value]
                ws[group_id_index] = created_group.id
            stats.add_synced(ws.title)
        except Exception as e:
            add_or_create_error(ws, sheet_value, e)
            stats.add_error(ws.title)
        finally:
            step_text = status_step_text(stats, ws.title)
            status.update(f"Syncing {ws.title}: {step_text}")

    return ws, id_mapping


sync_agreement_parameters = partial(sync_parameters, "Agreement")
sync_item_parameters = partial(sync_parameters, "Item")
sync_request_parameters = partial(sync_parameters, "Request")
sync_subscription_parameters = partial(sync_parameters, "Subscription")


def sync_items(
    mpt_client: MPTClient,
    file_handler: ExcelFileHandler,
    product: Product,
    values: SheetValueGenerator,
    items_groups_mapping: dict[str, ItemGroup],
    item_parameters_id_mapping: dict[str, Parameter],
    stats: ProductStatsCollector,
    status: Status,
) -> SyncResult[Item]:
    """
    Sync parameters by scope
    Returns mapping if ids {"<excel-id>": Item}
    """
    id_mapping = {}
    ws = file_handler._get_worksheet(constants.TAB_ITEMS)
    for sheet_value in values:
        try:
            sheet_value = setup_unit_of_measure(mpt_client, sheet_value)
            item = mpt_create_item(
                mpt_client,
                to_item_sync_json(
                    product.id,
                    items_groups_mapping,
                    item_parameters_id_mapping,
                    sheet_value,
                ),
            )

            id_index, _, sheet_id_value = find_value_for(
                constants.ID_COLUMN_NAME, sheet_value
            )

            # TODO: refactor, should be done out of this function
            ws[id_index] = item.id
            id_mapping[sheet_id_value] = item

            uom_id_index, _, sheet_id_value = find_value_for(
                constants.ITEMS_UNIT_ID, sheet_value
            )
            ws[uom_id_index] = sheet_id_value

            group_id_index, _, sheet_id_value = find_value_for(
                constants.ITEMS_GROUP_ID, sheet_value
            )
            created_group = items_groups_mapping[sheet_id_value]
            ws[group_id_index] = created_group.id
            stats.add_synced(ws.title)
        except Exception as e:
            add_or_create_error(ws, sheet_value, e)
            stats.add_error(ws.title)
        finally:
            step_text = status_step_text(stats, ws.title)
            status.update(f"Syncing {ws.title}: {step_text}")

    return ws, id_mapping


def create_item(
    mpt_client: MPTClient,
    active_account: Account,
    product_id: str,
    ws: Worksheet,
    sheet_value: list[SheetValue],
) -> None:
    sheet_value = setup_unit_of_measure(mpt_client, sheet_value)
    item = mpt_create_item(
        mpt_client,
        to_item_update_or_create_json(
            product_id,
            sheet_value,
            active_account.type == "Operations",
            ),
    )

    id_index, _, sheet_id_value = find_value_for(constants.ID_COLUMN_NAME, sheet_value)

    # TODO: refactor, should be done out of this function
    ws[id_index] = item.id

    uom_id_index, _, sheet_id_value = find_value_for(
        constants.ITEMS_UNIT_ID, sheet_value
    )
    ws[uom_id_index] = sheet_id_value


def setup_unit_of_measure(
    mpt_client: MPTClient, values: list[SheetValue]
) -> list[SheetValue]:
    unit_name = find_value_for(constants.ITEMS_UNIT_NAME, values)[2]

    uom = search_uom_by_name(mpt_client, unit_name)
    return set_value(constants.ITEMS_UNIT_ID, values, uom.id)


def sync_templates(
    mpt_client: MPTClient,
    file_handler: ExcelFileHandler,
    product: Product,
    values: SheetValueGenerator,
    all_parameters_id_mapping: dict[str, Parameter],
    stats: ProductStatsCollector,
    status: Status,
) -> SyncResult[Template]:
    """
    Sync templates
    Returns mapping if ids {"<excel-id>": Template}
    """
    id_mapping = {}
    ws = file_handler._get_worksheet(constants.TAB_TEMPLATES)
    for sheet_value in values:
        try:
            sheet_value = replace_parameter_variables(
                sheet_value, all_parameters_id_mapping
            )
            template = create_template(
                mpt_client,
                product,
                to_template_json(sheet_value),
            )

            content_index, _, sheet_content_value = find_value_for(
                constants.TEMPLATES_CONTENT, sheet_value
            )
            ws[content_index] = sheet_content_value

            id_index, _, sheet_id_value = find_value_for(
                constants.ID_COLUMN_NAME, sheet_value
            )

            # TODO: refactor, should be done out of this function
            ws[id_index] = template.id
            id_mapping[sheet_id_value] = template
            stats.add_synced(ws.title)
        except Exception as e:
            add_or_create_error(ws, sheet_value, e)
            stats.add_error(ws.title)
        finally:
            step_text = status_step_text(stats, ws.title)
            status.update(f"Syncing {ws.title}: {step_text}")

    return ws, id_mapping


def replace_parameter_variables(
    values: list[SheetValue], parameter_mapping: dict[str, Parameter]
) -> list[SheetValue]:
    _, _, content = find_value_for(constants.TEMPLATES_CONTENT, values)

    for sheet_param_id, param in parameter_mapping.items():
        content = content.replace(sheet_param_id, param.id)

    return set_value(constants.TEMPLATES_CONTENT, values, content)


def check_product_exists(
    mpt_client: MPTClient, product_definition_path: Path
) -> Product | None:
    file_handler = ExcelFileHandler(product_definition_path)
    general_values = file_handler.get_data_from_vertical_sheet(
        constants.TAB_GENERAL, constants.GENERAL_FIELDS
    )
    product_id = general_values[constants.GENERAL_PRODUCT_ID]["value"]

    _, products = get_products(mpt_client, 1, 0, query=f"id={product_id}")
    return products[0] if products else None
