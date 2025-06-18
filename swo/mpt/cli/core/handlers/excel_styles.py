from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill

general_tab_title_style = NamedStyle(name="general_tab_title")
general_tab_title_style.fill = PatternFill(fill_type="solid", fgColor="FF757171")
general_tab_title_style.font = Font(color="FFFFFF", bold=True)
general_tab_title_style.alignment = Alignment(horizontal="center", vertical="center")


horizontal_tab_style = NamedStyle(name="price_items_tab_style")
horizontal_tab_style.fill = PatternFill(fill_type="solid", fgColor="FF757171")
horizontal_tab_style.font = Font(color="FFFFFF", bold=True)
horizontal_tab_style.alignment = Alignment(horizontal="center", vertical="center")

number_format_style = NamedStyle(name="number_format_style")


def get_number_format_style(currency: str, precision: int) -> NamedStyle:
    currency_letter = {"USD": "$", "EUR": "€"}
    number_format_style.number_format = f"{currency_letter.get(currency, '')}#.{'0' * precision}"
    return number_format_style
