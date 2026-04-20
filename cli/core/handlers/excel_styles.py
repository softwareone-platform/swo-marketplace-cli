from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill

general_tab_title_style = NamedStyle(name="general_tab_title")
general_tab_title_style.fill = PatternFill(fill_type="solid", fgColor="FF757171")
general_tab_title_style.font = Font(color="FFFFFF", bold=True)
general_tab_title_style.alignment = Alignment(horizontal="center", vertical="center")


horizontal_tab_style = NamedStyle(name="price_items_tab_style")
horizontal_tab_style.fill = PatternFill(fill_type="solid", fgColor="FF757171")
horizontal_tab_style.font = Font(color="FFFFFF", bold=True)
horizontal_tab_style.alignment = Alignment(horizontal="center", vertical="center")

number_format_style = NamedStyle(name="number_format_style")


def get_number_format_style(currency: str, precision: int) -> NamedStyle:
    """Return a NamedStyle with a number format for the given currency and precision.

    Args:
        currency: The currency code (e.g., 'USD', 'EUR').
        precision: The number of decimal places.

    Returns:
        A NamedStyle with the specified number format.

    """
    currency_letter = {"USD": "$", "EUR": "€"}
    precision_format = "0" * precision
    currency_symbol = currency_letter.get(currency, "")
    number_format_style.number_format = f"{currency_symbol}#.{precision_format}"
    return number_format_style
