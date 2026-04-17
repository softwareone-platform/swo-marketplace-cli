import importlib
import re
from collections.abc import Generator
from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
from cli.core.handlers.errors import (
    RequiredFieldsError,
    RequiredFieldValuesError,
    RequiredSheetsError,
)
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

type SheetData = dict[str, Any]
type SheetDataGenerator = Generator[SheetData, None, None]

if TYPE_CHECKING:
    from cli.core.handlers.excel_file_handler import CellPosition


class ExcelWorkbookMixin:
    """Provide workbook lifecycle helpers."""

    _workbook_cache: Workbook | None
    file_path: Path
    _worksheets_cache: dict[str, Worksheet]

    @property
    def sheet_names(self) -> list[str]:
        return self.workbook.sheetnames

    @property
    def workbook(self) -> Workbook:
        if self._workbook_cache is None:
            excel_file_handler = importlib.import_module("cli.core.handlers.excel_file_handler")
            self._workbook_cache = excel_file_handler.load_workbook(self.file_path)

        return self._workbook_cache

    @classmethod
    def normalize_file_path(cls, file_path: str) -> Path:
        """Converts a file path string to a Path object with .xlsx extension."""
        return Path(file_path).with_suffix(".xlsx")

    def create(self):
        """Creates a new Excel workbook and saves it to the file_path."""
        wb = openpyxl.Workbook()
        wb.active.title = "General"
        wb.save(self.file_path)
        wb.close()
        self._workbook_cache = None
        self._worksheets_cache = {}


class ExcelValidationMixin:
    """Provide validation helpers for Excel sheets."""

    _get_fields_from_vertical_worksheet: Any
    get_data_from_vertical_sheet: Any
    _get_fields_from_horizontal_worksheet: Any

    @property
    def sheet_names(self) -> list[str]:
        raise NotImplementedError

    def check_required_sheet(self, required_sheets: tuple[str, ...]) -> None:
        """Checks if specified required sheets exist in the workbook."""
        missed_sheets = [sheet for sheet in required_sheets if sheet not in self.sheet_names]
        if missed_sheets:
            raise RequiredSheetsError(details=missed_sheets)

    def check_required_fields_in_vertical_sheet(
        self, sheet_name: str, required_fields: tuple[str, ...]
    ) -> None:
        """Checks if all required fields are present in a vertical sheet."""
        sheet_fields = self._get_fields_from_vertical_worksheet(sheet_name, 1)
        fields_missed = [field for field in required_fields if field not in sheet_fields]
        if fields_missed:
            raise RequiredFieldsError(details=fields_missed)

    def check_required_field_values_in_vertical_sheet(
        self, sheet_name: str, required_fields: list[str]
    ) -> None:
        """Checks if all required fields have values in a vertical sheet."""
        sheet_data = self.get_data_from_vertical_sheet(sheet_name)
        missed_values = []
        for required_field in required_fields:
            try:
                required_value = sheet_data[required_field]["value"]
            except KeyError:
                continue

            if required_value is None:
                missed_values.append(required_field)

        if missed_values:
            raise RequiredFieldValuesError(details=missed_values)

    def check_required_fields_in_horizontal_sheet(
        self, sheet_name: str, required_fields: list[str]
    ) -> None:
        """Checks if all required fields are present in a horizontal sheet."""
        col_fields = self._get_fields_from_horizontal_worksheet(sheet_name, 1)
        missed_fields = [
            required_field for required_field in required_fields if required_field not in col_fields
        ]
        if missed_fields:
            raise RequiredFieldsError(details=missed_fields)


class ExcelReadMixin:
    """Provide read helpers for Excel sheets."""

    _get_worksheet: Any
    _get_fields_from_horizontal_worksheet: Any

    def get_cell_value_by_coordinate(self, sheet_name: str, coordinate: str) -> str:
        """Retrieves the value of a specific cell in a sheet by its coordinate."""
        return self._get_worksheet(sheet_name)[coordinate].value

    def get_data_from_horizontal_sheet(
        self, sheet_name: str, fields: tuple[str, ...] | None = None
    ) -> SheetDataGenerator:
        """Retrieves data from a horizontally oriented sheet."""
        sheet = self._get_worksheet(sheet_name)
        header_fields = self._get_fields_from_horizontal_worksheet(sheet_name, 1)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                continue

            yield {
                header_fields[index]: {"value": cell.value, "coordinate": cell.coordinate}
                for index, cell in enumerate(row)
                if fields is None or header_fields[index] in fields
            }

    def get_data_from_vertical_sheet(
        self, sheet_name: str, fields: tuple[str, ...] | None = None
    ) -> SheetData:
        """Extracts data from a vertical sheet where the first column contains field names."""
        sheet = self._get_worksheet(sheet_name)
        sheet_iter = sheet.iter_rows(min_row=2)
        return {
            str(row[0].value): {"value": row[1].value, "coordinate": row[1].coordinate}
            for row in sheet_iter
            if self._is_non_empty_field_value(row[0].value)
            and (fields is None or row[0].value in fields)
        }

    def get_sheet_next_column(self, sheet_name: str) -> str:
        """Get the next available column letter in the specified sheet."""
        return get_column_letter(self._get_worksheet(sheet_name).max_column + 1)

    def get_sheet_next_row(self, sheet_name: str) -> int:
        """Get the next available row number in the specified sheet."""
        return self._get_worksheet(sheet_name).max_row + 1

    def get_values_for_dynamic_sheet(
        self, sheet_name: str, fields: tuple[str, ...], patterns: list[re.Pattern[str]]
    ) -> SheetDataGenerator:
        """Extracts data from a sheet with a dynamic column structure."""
        ws = self._get_worksheet(sheet_name)
        column_map = {}
        for header_index, column in enumerate(ws["1"]):
            if column.value and (
                column.value in fields
                or any(re.match(pattern, column.value) for pattern in patterns)
            ):
                column_map[header_index] = column.value

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if all(cell.value is None for cell in row):
                continue

            yield {
                column_name: {
                    "value": row[column_index].value,
                    "coordinate": f"{row[column_index].column_letter}{row[column_index].row}",  # type: ignore[union-attr]
                }
                for column_index, column_name in column_map.items()
            }

    def _is_non_empty_field_value(self, field_value: Any) -> bool:
        if field_value is None:
            return False

        return not isinstance(field_value, str) or bool(field_value.strip())


class ExcelWriteMixin:
    """Provide write helpers for Excel sheets."""

    _worksheets_cache: dict[str, Worksheet]
    file_path: Path
    _get_worksheet: Any

    @property
    def workbook(self) -> Workbook:
        raise NotImplementedError

    def merge_cells(self, sheet_name: str, range_string: str) -> None:
        """Merges a range of cells in the specified sheet."""
        self.workbook[sheet_name].merge_cells(range_string)

    def read(self) -> list[Any]:
        """Return the default empty payload for unsupported bulk workbook reads.

        Subclasses can override this when they need a workbook-wide read contract.
        """
        return []

    def save(self) -> None:
        """Saves the current workbook to the file path and cleans worksheet cache."""
        self.workbook.save(self.file_path)
        self._clean_worksheets()

    def write(self, sheet_rows: list[SheetData]) -> None:
        """Writes data to the Excel workbook."""
        for sheet in sheet_rows:
            for sheet_name, cells in sheet.items():
                try:
                    worksheet = self._get_worksheet(sheet_name)
                except KeyError:
                    worksheet = self.workbook.create_sheet(title=sheet_name)

                for coordinate, cell_value in cells.items():
                    worksheet[coordinate] = cell_value

        self.save()

    def write_cell(
        self,
        sheet_name: str,
        position: "CellPosition",
        cell_value: str,
        data_validation: DataValidation | None = None,
        style: NamedStyle | None = None,
    ) -> None:
        """Writes a value to a cell, applying style and data validation if provided."""
        try:
            sheet = self._get_worksheet(sheet_name)
        except KeyError:
            sheet = self.workbook.create_sheet(title=sheet_name)

        coordinate = f"{get_column_letter(position.col)}{position.row}"
        if style is not None:
            sheet[coordinate].style = style

        if data_validation is not None:
            if data_validation not in sheet.data_validations:
                sheet.add_data_validation(data_validation)
            data_validation.add(sheet[coordinate])

        sheet[coordinate] = cell_value

    def _clean_worksheets(self, sheet_name: str | None = None) -> None:
        if sheet_name is None:
            self._worksheets_cache = {}
            return

        self._worksheets_cache.pop(sheet_name, None)


class ExcelWorksheetMixin:
    """Provide worksheet access helpers."""

    _worksheets_cache: dict[str, Worksheet]

    @property
    def workbook(self) -> Workbook:
        raise NotImplementedError

    def _get_fields_from_horizontal_worksheet(self, worksheet_name: str, max_row: int) -> list[str]:
        return list(
            next(self._get_worksheet(worksheet_name).iter_rows(max_row=max_row, values_only=True))  # type: ignore[arg-type]
        )

    def _get_fields_from_vertical_worksheet(self, worksheet_name: str, max_col: int) -> list[str]:
        return list(
            next(self._get_worksheet(worksheet_name).iter_cols(max_col=max_col, values_only=True))  # type: ignore[arg-type]
        )

    def _get_worksheet(self, sheet_name: str) -> Worksheet:
        if self._worksheets_cache.get(sheet_name) is None:
            self._worksheets_cache[sheet_name] = self.workbook[sheet_name]

        return self._worksheets_cache[sheet_name]


class ExcelAccessMixin(ExcelWorkbookMixin, ExcelWorksheetMixin):
    """Group workbook and worksheet access helpers."""


class ExcelSheetMixin(ExcelValidationMixin, ExcelReadMixin, ExcelWriteMixin):
    """Group sheet validation, read, and write helpers."""
