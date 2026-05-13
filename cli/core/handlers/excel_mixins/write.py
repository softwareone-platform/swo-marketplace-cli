from pathlib import Path
from typing import TYPE_CHECKING, Any

from cli.core.handlers.excel_mixins.types import SheetData
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from cli.core.handlers.excel_file_handler import CellPosition


class ExcelWriteMixin:  # noqa: WPS214
    """Provide write helpers for Excel sheets."""

    _get_worksheet: Any
    _worksheets_cache: dict[str, Worksheet]
    file_path: Path

    @property
    def workbook(self) -> Workbook:
        raise NotImplementedError

    def merge_cells(self, sheet_name: str, range_string: str) -> None:
        """Merges a range of cells in the specified sheet."""
        self.workbook[sheet_name].merge_cells(range_string)

    def read(self) -> list[Any]:
        """Return the default empty payload for unsupported bulk workbook reads.

        Subclasses can override this when they need a workbook-wide read contract.
        """
        return []

    def save(self) -> None:
        """Saves the current workbook to the file path and cleans worksheet cache."""
        self.workbook.save(self.file_path)
        self._clean_worksheets()

    def write(self, sheet_rows: list[SheetData]) -> None:
        """Writes data to the Excel workbook."""
        for sheet in sheet_rows:
            for sheet_name, cells in sheet.items():
                self._write_cells(sheet_name, cells)
        self.save()

    def write_cell(
        self,
        sheet_name: str,
        position: "CellPosition",
        cell_value: str,
        data_validation: DataValidation | None = None,
        style: NamedStyle | None = None,
    ) -> None:
        """Writes a value to a cell, applying style and data validation if provided."""
        try:
            sheet = self._get_worksheet(sheet_name)
        except KeyError:
            sheet = self.workbook.create_sheet(title=sheet_name)

        coordinate = f"{get_column_letter(position.col)}{position.row}"
        if style is not None:
            sheet[coordinate].style = style

        if data_validation is not None:
            if data_validation not in sheet.data_validations:
                sheet.add_data_validation(data_validation)
            data_validation.add(sheet[coordinate])

        sheet[coordinate] = cell_value

    def _clean_worksheets(self, sheet_name: str | None = None) -> None:
        if sheet_name is None:
            self._worksheets_cache = {}
            return

        self._worksheets_cache.pop(sheet_name, None)

    def _write_cells(self, sheet_name: str, cells: dict) -> None:
        try:
            worksheet = self._get_worksheet(sheet_name)
        except KeyError:
            worksheet = self.workbook.create_sheet(title=sheet_name)
        for coordinate, cell_value in cells.items():
            worksheet[coordinate] = cell_value
