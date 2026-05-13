from pathlib import Path

import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelWorkbookMixin:
    """Provide workbook lifecycle helpers."""

    _workbook_cache: Workbook | None
    _worksheets_cache: dict[str, Worksheet]
    file_path: Path

    @property
    def sheet_names(self) -> list[str]:
        return self.workbook.sheetnames

    @property
    def workbook(self) -> Workbook:
        if self._workbook_cache is None:
            self._workbook_cache = load_workbook(self.file_path)

        return self._workbook_cache

    @classmethod
    def normalize_file_path(cls, file_path: str) -> Path:
        """Converts a file path string to a Path object with .xlsx extension."""
        return Path(file_path).with_suffix(".xlsx")

    def create(self):
        """Creates a new Excel workbook and saves it to the file_path."""
        wb = openpyxl.Workbook()
        wb.active.title = "General"
        wb.save(self.file_path)
        wb.close()
        self._workbook_cache = None
        self._worksheets_cache = {}
