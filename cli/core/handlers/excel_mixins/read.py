import re
from typing import Any

from cli.core.handlers.excel_mixins.types import (
    ColumnPatterns,
    SheetData,
    SheetDataGenerator,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReadMixin:  # noqa: WPS214
    """Provide read helpers for Excel sheets."""

    _get_fields_from_horizontal_worksheet: Any
    _get_worksheet: Any

    def get_cell_value_by_coordinate(self, sheet_name: str, coordinate: str) -> str:
        """Retrieves the value of a specific cell in a sheet by its coordinate."""
        return self._get_worksheet(sheet_name)[coordinate].value

    def get_data_from_horizontal_sheet(
        self, sheet_name: str, fields: tuple[str, ...] | None = None
    ) -> SheetDataGenerator:
        """Retrieves data from a horizontally oriented sheet."""
        sheet = self._get_worksheet(sheet_name)
        header_fields = self._get_fields_from_horizontal_worksheet(sheet_name, 1)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                continue

            yield {
                header_fields[index]: {"value": cell.value, "coordinate": cell.coordinate}
                for index, cell in enumerate(row)
                if fields is None or header_fields[index] in fields
            }

    def get_data_from_vertical_sheet(
        self, sheet_name: str, fields: tuple[str, ...] | None = None
    ) -> SheetData:
        """Extracts data from a vertical sheet where the first column contains field names."""
        result: SheetData = {}
        for field_cell, value_cell, *_ in self._get_worksheet(sheet_name).iter_rows(min_row=2):
            if not self._is_non_empty_field_value(field_cell.value):
                continue
            if fields is not None and field_cell.value not in fields:
                continue
            result[str(field_cell.value)] = {
                "value": value_cell.value,
                "coordinate": value_cell.coordinate,
            }
        return result

    def get_sheet_next_column(self, sheet_name: str) -> str:
        """Get the next available column letter in the specified sheet."""
        return get_column_letter(self._get_worksheet(sheet_name).max_column + 1)

    def get_sheet_next_row(self, sheet_name: str) -> int:
        """Get the next available row number in the specified sheet."""
        return self._get_worksheet(sheet_name).max_row + 1

    def get_values_for_dynamic_sheet(
        self, sheet_name: str, fields: tuple[str, ...], patterns: ColumnPatterns
    ) -> SheetDataGenerator:
        """Extracts data from a sheet with a dynamic column structure."""
        ws = self._get_worksheet(sheet_name)
        column_map = self._build_dynamic_column_map(ws, fields, patterns)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if all(cell.value is None for cell in row):
                continue

            yield {
                column_name: {
                    "value": row[column_index].value,
                    "coordinate": f"{row[column_index].column_letter}{row[column_index].row}",  # type: ignore[union-attr]
                }
                for column_index, column_name in column_map.items()
            }

    def _build_dynamic_column_map(
        self, ws: Worksheet, fields: tuple[str, ...], patterns: ColumnPatterns
    ) -> dict[int, str]:
        return {
            header_index: column.value
            for header_index, column in enumerate(ws["1"])
            if self._column_matches(column.value, fields, patterns)
        }

    def _column_matches(
        self, column_value: object, fields: tuple[str, ...], patterns: ColumnPatterns
    ) -> bool:
        if not column_value:
            return False
        if not isinstance(column_value, str):
            return False
        if column_value in fields:
            return True
        return any(re.match(pattern, column_value) for pattern in patterns)

    def _is_non_empty_field_value(self, field_value: Any) -> bool:
        if field_value is None:
            return False

        return not isinstance(field_value, str) or bool(field_value.strip())
